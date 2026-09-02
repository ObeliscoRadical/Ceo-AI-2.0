from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import HTTPException, Request, Response, Depends, Header, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from typing import List, Optional, Dict, Any
from bson import ObjectId
from datetime import datetime, timezone, timedelta
import logging, uuid, jwt, bcrypt, io, json, requests, random, stripe, httpx, hashlib, secrets, base64, asyncio, threading, time, re, urllib.parse
from google import genai
from google.genai import types

# ---------------------------------------------------------------- message / event classes
class TextDelta:
    def __init__(self, content: str):
        self.content = content

class StreamDone:
    pass

class UserMessage:
    def __init__(self, text: str, file_contents: list = None):
        self.text = text
        self.file_contents = file_contents or []

class ImageContent:
    def __init__(self, image_base64: str):
        self.image_base64 = image_base64

class FileContentWithMimeType:
    def __init__(self, file_path: str, mime_type: str = "application/pdf"):
        self.file_path = file_path
        self.mime_type = mime_type

# ---------------------------------------------------------------- config
JWT_ALGORITHM = "HS256"
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
genai_client = genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None

# LiteLLM Central Gateway Configuration (OpenAI Compatible)
LITELLM_BASE_URL = (
    os.environ.get("LITELLM_BASE_URL")
    or os.environ.get("OPENAI_BASE_URL")
    or os.environ.get("BASE_URL")
    or "https://berriailitellm-databasev1826rc3-production-2f93.up.railway.app"
).rstrip("/")
LITELLM_API_KEY = (
    os.environ.get("LITELLM_API_KEY")
    or os.environ.get("OPENAI_API_KEY")
    or os.environ.get("API_KEY")
    or ""
)
LITELLM_MODEL = (
    os.environ.get("LITELLM_MODEL")
    or os.environ.get("MODEL")
    or "Gemini Auto Key"
)

APP_NAME = "ceo-ai-2.0"
stripe.api_key = os.environ.get("STRIPE_SECRET_KEY") or "sk_test_stripe"
STRIPE_WEBHOOK_SECRET = os.environ.get("STRIPE_WEBHOOK_SECRET", "")
EMAIL_FROM_NAME = os.environ.get("EMAIL_FROM_NAME", "CEO AI 2.0")

MODEL_MAP = {
    "claude": ("gemini", "gemini-3.7-flash"),
    "gpt": ("gemini", "gemini-3.7-flash"),
    "gemini": ("gemini", "gemini-3.7-flash"),
}
CURRENCY_SYMBOL = {"EUR": "€", "BRL": "R$", "USD": "$"}

mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017/ceo_ai_2_0')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'ceo_ai_2_0')]

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------- auth helpers
def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def set_auth_cookie(response: Response, token: str):
    is_secure = os.environ.get("COOKIE_SECURE", "false").lower() in ("true", "1")
    samesite = "none" if is_secure else "lax"
    response.set_cookie(key="access_token", value=token, httponly=True, secure=is_secure,
                        samesite=samesite, max_age=604800, path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Não autenticado")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilizador não encontrado")
        user["id"] = str(user["_id"])
        user["is_premium"] = bool(user.get("is_premium"))
        user.pop("_id", None)
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Sessão expirada")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

async def is_premium(user_id: str) -> bool:
    u = await db.users.find_one({"_id": ObjectId(user_id)})
    return bool(u and u.get("is_premium"))

# ---------------------------------------------------------------- company resolution
async def resolve_company(user_id: str, company_id: Optional[str] = None):
    if company_id:
        c = await db.companies.find_one({"_id": ObjectId(company_id), "user_id": user_id})
        if c:
            return c
    s = await db.settings.find_one({"user_id": user_id}) or {}
    acid = s.get("active_company_id")
    if acid:
        c = await db.companies.find_one({"_id": ObjectId(acid), "user_id": user_id})
        if c:
            return c
    return await db.companies.find_one({"user_id": user_id})

async def active_company_id(user_id: str) -> Optional[str]:
    c = await resolve_company(user_id)
    if not c:
        return None
    cid = str(c["_id"])
    # migrate orphan entries to the active company (one-time, cheap)
    await db.entries.update_many({"user_id": user_id, "company_id": {"$exists": False}},
                                 {"$set": {"company_id": cid}})
    return cid

# ---------------------------------------------------------------- storage
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

def init_storage():
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    return True

def put_object(path: str, data: bytes, content_type: str) -> dict:
    target_path = UPLOAD_DIR / path
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with open(target_path, "wb") as f:
        f.write(data)
    return {"path": path, "url": f"/uploads/{path}"}

async def generate_post_visual_scenes(titulo: str, legenda: str, hook: str = "", product_name: str = "", sector: str = "", company_name: str = "") -> list[str]:
    """Usa o Gemini como Diretor Criativo Fotográfico para criar 3 cenas visuais em inglês ultra-contextualizadas com o GANCHO (HOOK), PRODUTO e HISTÓRIA real da peça."""
    system = (
        "You are an award-winning Commercial Photography Creative Director for high-converting advertising. "
        "Your job is to translate a social media post's HOOK, TITLE, PRODUCT, and CAPTION into 3 distinct, "
        "ultra-concrete, photorealistic commercial photography scene prompts for the Flux AI image generator.\n\n"
        "MANDATORY RULES:\n"
        "1. NO ABSTRACT CONCEPTS, NO EMPTY CORRIDORS, NO FLOATING OBJECTS, NO GENERIC LAMPS/WALLS.\n"
        "2. The image MUST depict real humans in action, authentic professional workplace/lifestyle situations that directly reflect the emotional tension of the HOOK and the value of the PRODUCT (e.g. focused business owner managing operations, skilled technician working on site, confident executive in modern office, customer experiencing the direct benefit).\n"
        "3. Every prompt must be in descriptive English (35-50 words), detailing: Subject, Action, Facial Expression, Setting/Environment, Lighting (e.g. warm cinematic natural light), Camera Shot (e.g. medium commercial portrait, sharp focus on subject), 8k resolution, photorealistic.\n"
        "4. Always append: 'award winning commercial photography, 8k, photorealistic, sharp focus, no text, no words, no letters, no logos, no watermark, no CGI, no cartoon'.\n"
        "5. Return ONLY a JSON list of 3 strings: [\"prompt1\", \"prompt2\", \"prompt3\"]."
    )
    user_prompt = (
        f"Empresa: {company_name or 'Empresa Líder'} (Setor: {sector or 'Serviços Especializados'})\n"
        f"Produto/Serviço: {product_name or 'Solução Profissional'}\n"
        f"Gancho (Hook de Abertura): {hook}\n"
        f"Título da Peça: {titulo}\n"
        f"Legenda / Copy: {legenda[:400]}\n\n"
        "Generate 3 distinct realistic commercial photography scene prompts strictly grounded in the HOOK and PRODUCT context in JSON format:"
    )
    try:
        res = await ai_text(system, user_prompt)
        start = res.find('[')
        end = res.rfind(']') + 1
        if start >= 0 and end > start:
            parsed = json.loads(res[start:end])
            if isinstance(parsed, list) and len(parsed) >= 1:
                return [str(p).strip() for p in parsed if p][:3]
    except Exception as e:
        logger.warning(f"generate_post_visual_scenes note: {e}")
    
    # Concrete context fallback
    subj = product_name or company_name or "commercial business"
    hook_desc = hook or titulo or "business management operations"
    return [
        f"Professional business owner actively working at a modern office desk with laptop and documents illustrating '{hook_desc}', warm cinematic sunlight, realistic human expression, sharp focus, award winning commercial photography, 8k, photorealistic, no text, no watermark",
        f"Close up dynamic shot of a specialist in uniform with modern tools delivering {subj}, clean professional facility, crisp lighting, depth of field, 8k photorealistic commercial photography, no text, no watermark",
        f"Confident executive reviewing strategic growth dashboard on tablet in a bright contemporary workplace, natural window lighting, 8k, editorial commercial photography, no text, no watermark"
    ]

async def search_topic_exact_images(query: str, count: int = 3) -> list[bytes]:
    """Procura imagens reais e profissionais no DuckDuckGo estritamente correspondentes ao tema."""
    results = []
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    }
    try:
        async with httpx.AsyncClient(headers=headers, timeout=10.0, follow_redirects=True) as client:
            token_res = await client.get("https://duckduckgo.com/", params={"q": f"{query} commercial photography"})
            vqd_match = re.search(r'vqd=([\d-]+)', token_res.text) or re.search(r'vqd="([^"]+)"', token_res.text)
            if vqd_match:
                vqd = vqd_match.group(1)
                search_res = await client.get("https://duckduckgo.com/i.js", params={"l": "us-en", "o": "json", "q": f"{query} professional", "vqd": vqd, "f": ",,,", "p": "1"})
                if search_res.status_code == 200:
                    items = search_res.json().get("results", [])
                    for it in items:
                        if len(results) >= count:
                            break
                        img_url = it.get("image")
                        if img_url and not img_url.endswith(".svg"):
                            try:
                                dl_res = await client.get(img_url, timeout=6.0)
                                if dl_res.status_code == 200 and len(dl_res.content) > 6000 and dl_res.headers.get("content-type", "").startswith("image/"):
                                    results.append(dl_res.content)
                            except Exception:
                                pass
    except Exception as e:
        logger.debug(f"search_topic_exact_images note: {e}")
    return results

async def generate_marketing_images(prompt: str = "", number_of_images: int = 3, scene_prompts: list[str] = None, topic_query: str = "") -> list[bytes]:
    """Gera 1..N imagens de marketing ultra-realistas com qualidade editorial fotográfica de topo."""
    count = max(1, min(int(number_of_images or 3), 4))
    
    # Prefixos fotográficos de alta fidelidade cinematográfica (estilo Sony A7R / Hasselblad)
    PHOTO_ENHANCERS = [
        "captured on 35mm Hasselblad H6D-100c, 85mm f/1.4 portrait lens, soft natural window lighting, rich skin textures, authentic expressions, hyper-detailed, award-winning editorial commercial photography, 8k uhd, cinematic color grading, photorealistic, no illustration, no 3d render, no anime, no text, no watermark",
        "captured on Sony A7R V, 50mm f/1.2 G Master lens, cinematic warm golden hour ambient lighting, subtle depth of field, authentic real-life environment, sharp focus on subject, commercial advertising photography, 8k, hyper-realistic, no cartoon, no CGI, no watermark, no text",
        "captured on Leica SL2, 24-70mm f/2.8 lens, modern architectural corporate lighting, crisp details, natural realistic shadows, depth, high dynamic range, hyperrealistic editorial photoshoot, 8k, pristine quality, no drawing, no CGI, no watermark, no text"
    ]
    
    if not scene_prompts:
        clean_p = prompt or "successful professional entrepreneur managing operations in high-end modern workspace"
        scene_prompts = [
            f"{clean_p}, {PHOTO_ENHANCERS[i % len(PHOTO_ENHANCERS)]}"
            for i in range(count)
        ]
    else:
        scene_prompts = [
            f"{sp}, {PHOTO_ENHANCERS[i % len(PHOTO_ENHANCERS)]}"
            for i, sp in enumerate(scene_prompts[:count])
        ]
    
    results = []
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}

    poll_timeout = httpx.Timeout(90.0, connect=15.0)
    async with httpx.AsyncClient(timeout=poll_timeout, follow_redirects=True, headers=headers) as client:
        for idx, scene in enumerate(scene_prompts[:count]):
            img_bytes = None
            base_seed = int(time.time() * 1000) % 1000000 + (idx * 337)
            enc_scene = urllib.parse.quote(scene[:420])

            # 1. Tentar Pollinations AI (Flux Pro / Flux Realism com seeds variadas)
            attempts = [("flux-realism", base_seed), ("flux", base_seed + 1), ("flux-pro", base_seed + 2), ("turbo", base_seed + 3)]
            for model_name, seed in attempts:
                poll_url = f"https://image.pollinations.ai/prompt/{enc_scene}?width=1080&height=1080&seed={seed}&nologo=true&model={model_name}&enhance=true"
                try:
                    res = await client.get(poll_url)
                    if res.status_code == 200 and len(res.content) > 5000 and not res.content.startswith(b'{"error"'):
                        img_bytes = res.content
                        break
                except Exception as e:
                    logger.debug(f"Pollinations model {model_name} note: {e}")

            if img_bytes:
                results.append(img_bytes)

            if len(results) >= count:
                break
            await asyncio.sleep(0.5)

    # 2. Fallback de Imagens Reais Exatamente Alinhadas ao Título/Tema
    if len(results) < count:
        needed = count - len(results)
        search_q = topic_query or (scene_prompts[0] if scene_prompts else "business management professional")
        clean_q = re.sub(r'(photorealistic|8k|no text|no watermark|no logos|no cgi|no abstract|commercial photography|,)', ' ', search_q)
        topic_images = await search_topic_exact_images(clean_q.strip(), count=needed)
        for t_img in topic_images:
            results.append(t_img)
            if len(results) >= count:
                break

    # 3. Fallback Procedural caso tudo falhe
    while len(results) < count:
        try:
            from PIL import Image, ImageDraw
            idx = len(results)
            palettes = [
                ((15, 23, 42), (30, 58, 138), (59, 130, 246)),
                ((15, 23, 42), (19, 78, 74), (16, 185, 129)),
                ((24, 24, 27), (88, 28, 135), (168, 85, 247)),
                ((15, 23, 42), (124, 45, 18), (249, 115, 22)),
            ]
            c_bg, c_mid, c_accent = palettes[idx % len(palettes)]
            img = Image.new("RGBA", (1080, 1080), c_bg)
            draw = ImageDraw.Draw(img)
            for r in range(500, 50, -10):
                alpha = int(35 * (1 - r / 500))
                draw.ellipse([540 - r, 540 - r, 540 + r, 540 + r], fill=(c_accent[0], c_accent[1], c_accent[2], alpha))
            draw.rectangle([60, 60, 1020, 1020], outline=c_accent, width=4)
            draw.line([(60, 540), (1020, 540)], fill=(c_mid[0], c_mid[1], c_mid[2], 120), width=2)
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            results.append(buf.getvalue())
        except Exception as e:
            logger.error(f"Procedural fallback error: {e}")
            break

    return results

async def generate_marketing_image(prompt: str) -> bytes:
    """Compat wrapper para chamadas que pedem 1 imagem."""
    imgs = await generate_marketing_images(prompt, number_of_images=1)
    return imgs[0]

def prepare_logo(data: bytes) -> bytes:
    """Normaliza o logo para PNG RGBA. Se não tiver transparência, remove o fundo claro
    ligado às bordas (mantém partes claras internas do logo, ex.: texto branco)."""
    from PIL import Image
    import numpy as np
    img = Image.open(io.BytesIO(data)).convert("RGBA")
    arr = np.array(img)
    if (arr[:, :, 3] < 250).mean() <= 0.02:
        try:
            from scipy import ndimage
            rgb = arr[:, :, :3].astype(int)
            mx, mn = rgb.max(2), rgb.min(2)
            light = (mn > 205) & ((mx - mn) < 30)
            labeled, _ = ndimage.label(light)
            border = (set(labeled[0, :]) | set(labeled[-1, :]) |
                      set(labeled[:, 0]) | set(labeled[:, -1]))
            border.discard(0)
            if border:
                arr[np.isin(labeled, list(border)), 3] = 0
                img = Image.fromarray(arr, "RGBA")
        except Exception as e:
            logger.error(f"prepare_logo keying: {e}")
    out = io.BytesIO(); img.save(out, format="PNG"); return out.getvalue()

def composite_logo(base_bytes: bytes, logo_bytes: bytes) -> bytes:
    """Sobrepõe o logo REAL da empresa (canto inferior direito) sobre a imagem gerada, mantendo-o nítido."""
    from PIL import Image
    base = Image.open(io.BytesIO(base_bytes)).convert("RGBA")
    logo = Image.open(io.BytesIO(logo_bytes)).convert("RGBA")
    bw, bh = base.size
    target_w = max(1, int(bw * 0.26))
    ratio = target_w / logo.width
    logo = logo.resize((target_w, max(1, int(logo.height * ratio))), Image.LANCZOS)
    margin = int(bw * 0.04)
    x, y = bw - logo.width - margin, bh - logo.height - margin
    base.alpha_composite(logo, (x, y))
    out = io.BytesIO()
    base.convert("RGB").save(out, format="PNG")
    return out.getvalue()

async def store_public_media(uid: str, data: bytes, ct: str = "image/png") -> str:
    """Guarda bytes de imagem e devolve URL público absoluto no backend."""
    mid = str(uuid.uuid4())
    await db.social_media.insert_one({
        "_id": mid, 
        "user_id": uid,
        "data": base64.b64encode(data).decode(), 
        "content_type": ct,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    backend_port = os.environ.get("PORT", "8001")
    return f"http://localhost:{backend_port}/api/public/media/{mid}"

def extract_document_text(data: bytes, content_type: str, filename: str) -> str:
    name = (filename or "").lower(); ct = (content_type or "").lower()
    try:
        if name.endswith(".pdf") or "pdf" in ct:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            return "\n".join((p.extract_text() or "") for p in reader.pages[:25])
        if name.endswith(".xlsx") or "sheet" in ct or "excel" in ct:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            out = []
            for ws in wb.worksheets[:6]:
                out.append(f"# {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        out.append("\t".join(cells))
                    if len(out) > 2500:
                        break
            return "\n".join(out)
        if name.endswith(".docx") or "wordprocessing" in ct:
            import docx
            d = docx.Document(io.BytesIO(data))
            return "\n".join(p.text for p in d.paragraphs)
        if name.endswith((".csv", ".txt", ".tsv")) or ct.startswith("text"):
            return data.decode("utf-8", errors="ignore")
    except Exception as e:
        logger.error(f"extract_document_text error: {e}")
    return ""

async def analyze_document(text: str, doc_type: str, filename: str) -> dict:
    if not text or len(text.strip()) < 20:
        return {"analysable": False, "relevant": False, "quality": "low",
                "summary": "Carregado, mas sem texto legível para análise automática (ex: imagem ou digitalização)."}
    prompt = (
        f"És um analista financeiro. Analisa este documento ('{filename}', categoria '{doc_type}') e devolve APENAS JSON: "
        '{"analysable":true,"relevant":bool,"quality":"high"|"medium"|"low","doc_kind":str,"period":str,"summary":str,'
        '"figures":{"revenue":number|null,"ebitda":number|null,"net_profit":number|null,"assets":number|null,'
        '"liabilities":number|null,"recurring_revenue":number|null,"currency":str|null},'
        '"contracts":{"count":number|null,"total_value":number|null,"recurring":bool}|null,'
        '"red_flags":[str],"strengths":[str]}. '
        "'relevant'=true se contém dados financeiros/contratuais úteis para avaliar a empresa. "
        "'quality': 'high' se são demonstrações/contratos formais com números claros; 'medium' se parcial; 'low' se pouco útil. "
        "Preenche 'figures' APENAS com números presentes no documento (senão null). 'summary' em 1-2 frases. "
        "Português europeu. Sem texto fora do JSON.\n\nCONTEÚDO:\n" + text[:8000]
    )
    ai = await ai_json("És um analista financeiro rigoroso. Respondes só com JSON.", prompt)
    return ai or {"analysable": True, "relevant": False, "quality": "low", "summary": "Não foi possível analisar o documento."}


FINANCE_MODEL = os.environ.get("FINANCE_MODEL", "gemini-2.5-pro")

TOTAL_KEYS = ["ativo_total", "ativo_nao_corrente", "ativo_corrente", "passivo_total", "capital_proprio",
              "vendas_e_servicos", "rendimentos_totais", "gastos_totais", "resultado_liquido", "ebitda"]

FIN_SCHEMA_PROMPT = (
    "És um Contabilista Certificado português (SNC). Lê este documento financeiro (balancete analítico, IES/DA, "
    "Modelo 22, declaração de IVA, demonstração de resultados ou balanço) e devolve APENAS JSON válido "
    "(sem markdown, sem texto fora do JSON).\n"
    "1) Identifica 'doc_type' (balancete|ies|modelo22|iva|demonstracao_resultados|balanco|outro) e 'year' (ano do exercício).\n"
    "2) SE o documento for uma DEMONSTRAÇÃO FINANCEIRA FORMAL (IES/DA, Balanço, Demonstração de Resultados, Modelo 22) "
    "que JÁ APRESENTA TOTAIS IMPRESSOS: LÊ esses valores oficiais diretamente (NÃO recalcules) e preenche 'totals'. "
    "Nota: nas demonstrações formais o 'capital_proprio' (Total do capital próprio) JÁ INCLUI o resultado líquido do período. "
    "Devolve 'lines' como lista vazia [].\n"
    "3) SE for um BALANCETE ANALÍTICO (lista de contas com saldos): deixa TODOS os campos de 'totals' a null e devolve "
    "em 'lines' as CONTAS DE RAZÃO (agregados de 2 dígitos: 11,12,13,21,22,23,24,25,26,27,28,31,32,33,41,42,43,44,45,"
    "51,55,56,58,59,61,62,63,64,65,68,69,71,72,75,78,79). NÃO devolvas subcontas de detalhe (ex: 2111001) — só os "
    "agregados — para nunca duplicares. Inclui a conta 81/88 (Resultado líquido do período). "
    "Cada linha: {\"code\":str,\"name\":str,\"balance\":number,\"nature\":str}, onde 'balance' é o saldo final "
    "(POSITIVO se devedor, NEGATIVO se credor) e 'nature' é um de: \"ativo_nc\",\"ativo_c\",\"passivo\",\"capital\","
    "\"gasto\",\"rendimento\",\"resultado\",\"outro\" (ativo_nc=classe 4; ativo_c=caixa/bancos/clientes/inventários/"
    "IVA a recuperar/outros devedores; passivo=fornecedores/financiamentos/Estado a pagar/outros credores; "
    "capital=classe 5; gasto=classe 6 e compras 31; rendimento=classe 7; resultado=81/88).\n"
    "'totals' = {\"ativo_total\":n|null,\"ativo_nao_corrente\":n|null,\"ativo_corrente\":n|null,\"passivo_total\":n|null,"
    "\"capital_proprio\":n|null,\"vendas_e_servicos\":n|null,\"rendimentos_totais\":n|null,\"gastos_totais\":n|null,"
    "\"resultado_liquido\":n|null,\"ebitda\":n|null} (ponto decimal, sem separador de milhar).\n"
    "NUNCA inventes números. Estrutura final: {\"doc_type\":str,\"year\":number|null,\"currency\":\"EUR\","
    "\"totals\":{...},\"lines\":[...],\"summary\":str}. Português europeu."
)

def _recon_from_totals(t):
    """Reconciliação de demonstrações formais (IES/Balanço): capital_proprio JÁ inclui o resultado."""
    a = t.get("ativo_total"); p = t.get("passivo_total"); c = t.get("capital_proprio")
    if isinstance(a, (int, float)) and isinstance(p, (int, float)) and isinstance(c, (int, float)):
        diff = round(a - (p + c), 2)
        tol = max(1000.0, abs(a) * 0.02)
        return (abs(diff) <= tol), diff
    return None, None

def _snc_reconcile(lines):
    """Soma determinística por natureza SNC. Dedup de contas-filhas por prefixo + netting por raiz de 2 dígitos
    (débito−crédito cancela IVA dedutível vs liquidado). Conta 81/88 = resultado autoritativo.
    Devolve (totals, reconciled, diff, kept_lines)."""
    clean = []
    for l in lines or []:
        if not isinstance(l, dict):
            continue
        code = str(l.get("code") or "").strip()
        bal = l.get("balance")
        if not code or not isinstance(bal, (int, float)):
            continue
        clean.append({"code": code, "name": l.get("name", ""), "balance": float(bal),
                      "nature": (l.get("nature") or "outro")})
    clean.sort(key=lambda l: len(l["code"]))
    kept = []; kept_codes = []
    for l in clean:
        if any(l["code"] != kc and l["code"].startswith(kc) for kc in kept_codes):
            continue  # descendente de uma conta já contabilizada
        kept.append(l); kept_codes.append(l["code"])

    anc = 0.0; anc_found = False        # ativo não corrente (classe 4, líquido de depreciações)
    cap = 0.0; cap_found = False        # capital próprio (classe 5)
    rend = 0.0; vendas = 0.0; rend_found = False
    gastos = 0.0; gastos_found = False
    resultado = None
    net_by_root = {}                    # ativo_c/passivo agregados e netted por raiz de 2 dígitos
    for l in kept:
        nat = l["nature"]; code = l["code"]; v = l["balance"]; root = code[:2]
        if nat == "ativo_nc":
            anc += v; anc_found = True
        elif nat == "capital":
            cap += (-v); cap_found = True   # crédito (negativo) → aumenta capital próprio
        elif nat == "rendimento":
            rend_found = True; rend += abs(v)
            if root in ("71", "72"):
                vendas += abs(v)
        elif nat == "gasto":
            gastos_found = True; gastos += abs(v)
        elif nat == "resultado" and abs(v) > 0.01:
            resultado = round(-v, 2)   # crédito (negativo) = lucro → positivo
        elif nat in ("ativo_c", "passivo"):
            net_by_root[root] = net_by_root.get(root, 0.0) + v

    bal_found = bool(net_by_root)
    ativo_c = 0.0; passivo = 0.0
    for _root, net in net_by_root.items():
        if net >= 0:
            ativo_c += net
        else:
            passivo += (-net)
    ativo_nc = round(anc, 2) if anc_found else None
    ativo_c = round(ativo_c, 2) if bal_found else None
    passivo = round(passivo, 2) if bal_found else None
    capital = round(cap, 2) if cap_found else None
    if resultado is None and rend_found and gastos_found:
        resultado = round(rend - gastos, 2)
    ativo_total = None
    if anc_found or bal_found:
        ativo_total = round((anc if anc_found else 0.0) + (ativo_c or 0.0), 2)
    totals = {
        "ativo_total": ativo_total, "ativo_nao_corrente": ativo_nc, "ativo_corrente": ativo_c,
        "passivo_total": passivo, "capital_proprio": capital,
        "vendas_e_servicos": (round(vendas, 2) if rend_found else None),
        "rendimentos_totais": (round(rend, 2) if rend_found else None),
        "gastos_totais": (round(gastos, 2) if gastos_found else None),
        "resultado_liquido": resultado, "ebitda": None,
    }
    reconciled = None; diff = None
    if ativo_total is not None and passivo is not None and capital is not None:
        equity = capital + (resultado or 0)
        diff = round(ativo_total - (passivo + equity), 2)
        tol = max(1000.0, ativo_total * 0.02)
        reconciled = abs(diff) <= tol
    return totals, reconciled, diff, kept

async def extract_financial_document(data: bytes, content_type: str, filename: str):
    import json as _json
    ct = (content_type or "").lower(); name = (filename or "").lower()
    sysmsg = "És um Contabilista Certificado português. Respondes só com JSON válido."
    
    parts = []
    if name.endswith(".pdf") or "pdf" in ct:
        parts.append(types.Part.from_bytes(data=data, mime_type="application/pdf"))
        prompt_text = FIN_SCHEMA_PROMPT
    elif ct.startswith("image/"):
        mime = "image/png" if "png" in ct else ("image/jpeg" if "jp" in ct else "image/webp")
        parts.append(types.Part.from_bytes(data=data, mime_type=mime))
        prompt_text = FIN_SCHEMA_PROMPT
    else:
        text_inline = extract_document_text(data, content_type, filename)[:15000]
        prompt_text = FIN_SCHEMA_PROMPT + "\n\nCONTEÚDO:\n" + (text_inline or "")
    parts.append(prompt_text)

    try:
        parsed = await ai_json(sysmsg, parts, model=FINANCE_MODEL)
        if not parsed:
            return None
        printed = parsed.get("totals") if isinstance(parsed.get("totals"), dict) else {}
        has_printed = any(isinstance(printed.get(k), (int, float)) and printed.get(k) not in (None, 0)
                          for k in ("ativo_total", "vendas_e_servicos", "resultado_liquido", "passivo_total"))
        if has_printed:
            t = {k: (printed.get(k) if isinstance(printed.get(k), (int, float)) else None) for k in TOTAL_KEYS}
            reconciled, diff = _recon_from_totals(t)
            parsed["totals"] = t
            parsed["reconciled"] = reconciled
            parsed["reconciliation_diff"] = diff
            parsed["lines"] = parsed.get("lines") or []
        else:
            totals, reconciled, diff, kept = _snc_reconcile(parsed.get("lines"))
            parsed["totals"] = totals
            parsed["reconciled"] = reconciled
            parsed["reconciliation_diff"] = diff
            parsed["lines"] = kept or parsed.get("lines")
        return parsed
    except Exception as e:
        logger.error(f"extract_financial_document error: {e}")
        return None


def rag(value, good, warn, reverse=False):
    if reverse:
        if value <= good: return "green"
        if value <= warn: return "amber"
        return "red"
    if value >= good: return "green"
    if value >= warn: return "amber"
    return "red"


def compute_balance(company: dict, profile: dict, entries_net: float = 0.0):
    """Single source of truth for the company balance sheet."""
    profile = profile or {}
    has = bool(profile)
    cash = float(profile.get("cash_balance", 0) or 0) if has else (float((company or {}).get("bank_balance", 0) or 0) + entries_net)
    assets_items = sum(float(a.get("amount", 0) or 0) for a in (profile.get("assets") or []))
    liab_items = sum(float(l.get("amount", 0) or 0) for l in (profile.get("liabilities") or []))
    debt = float(profile.get("total_debt", 0) or 0)
    total_assets = cash + assets_items
    total_liabilities = debt + liab_items
    return {"cash": round(cash, 2), "total_assets": round(total_assets, 2),
            "total_liabilities": round(total_liabilities, 2),
            "net_worth": round(total_assets - total_liabilities, 2)}

def compute_valuation(profile: dict, bal: dict):
    """Estimated company value = patrimonial floor (net worth) + earnings goodwill.
    Honest small-business model: assets/liabilities base plus a multiple on positive annual profit."""
    profile = profile or {}
    net_worth = bal.get("net_worth", 0)
    revenue = float(profile.get("monthly_revenue", 0) or 0)
    fixed = sum(float(c.get("amount", 0) or 0) for c in (profile.get("fixed_costs") or []))
    var_pct = max(0.0, min(100.0, float(profile.get("variable_costs_pct", 0) or 0)))
    profit_m = revenue - fixed - revenue * var_pct / 100.0
    annual_profit = profit_m * 12.0
    margin = (profit_m / revenue * 100.0) if revenue > 0 else 0.0
    mult = 0.0
    if annual_profit > 0:
        mult = 2.0
        if margin >= 10: mult += 0.5
        if margin >= 20: mult += 0.5
        if margin >= 30: mult += 0.5
    goodwill = max(0.0, annual_profit) * mult
    floor = net_worth if net_worth > 0 else 0.0
    value = round(max(floor + goodwill, bal.get("cash", 0) or 0), 2)
    method = "patrimonial + rendimento" if goodwill > 0 else "patrimonial"
    return {"value": value, "net_worth": round(net_worth, 2),
            "annual_profit": round(annual_profit, 2), "multiple": mult,
            "goodwill": round(goodwill, 2), "method": method}


async def compute_confidence(user_id: str, has_profile: bool):
    docs = await db.documents.find({"user_id": user_id, "is_deleted": False}).to_list(500)
    figs, verified = {}, 0
    for d in docs:
        a = d.get("analysis") or {}
        if a.get("relevant") and a.get("quality") in ("high", "medium"):
            verified += 1
        for k, v in (a.get("figures") or {}).items():
            if isinstance(v, (int, float)) and v and k not in figs:
                figs[k] = v
    has_fin = any(figs.get(k) for k in ("revenue", "ebitda", "net_profit", "assets", "liabilities"))
    if has_profile and has_fin and verified >= 1:
        tier, margin = "Avaliação Fundamentada", 0.12
    elif has_profile or has_fin:
        tier, margin = "Estimativa Fundamentada", 0.20
    else:
        tier, margin = "Estimativa Inteligente", 0.35
    return {"tier": tier, "margin": margin, "based_on_documents": has_fin,
            "documents_analyzed": verified, "figures": figs}


FORMAL_ORDER = {"ies": 0, "demonstracao_resultados": 1, "balanco": 2, "balancete": 3, "modelo22": 4, "iva": 5, "outro": 6}
DOC_LABEL = {"ies": "IES", "demonstracao_resultados": "Demonstração de Resultados", "balanco": "Balanço",
             "balancete": "Balancete", "modelo22": "Modelo 22", "iva": "Declaração de IVA", "outro": "Documento"}

async def latest_official_financials(user_id: str, cid):
    """Números do documento oficial mais recente (IES/balancete/DR...) para a empresa ativa.
    Fonte autoritativa do valor da empresa. Prefere o ano mais recente e, no mesmo ano, o documento mais formal."""
    if not cid:
        return None
    rows = await db.financial_extractions.find({"user_id": user_id, "company_id": cid}).to_list(50)
    rows = [r for r in rows if r.get("year")]
    if not rows:
        return None
    top_year = max(int(r["year"]) for r in rows)
    same = [r for r in rows if int(r["year"]) == top_year]
    same.sort(key=lambda r: FORMAL_ORDER.get(r.get("doc_type", "outro"), 9))
    r = same[0]
    dt = r.get("doc_type", "outro")
    assets = r.get("ativo_total"); liab = r.get("passivo_total"); equity = r.get("capital_proprio")
    revenue = r.get("vendas_e_servicos"); profit = r.get("resultado_liquido"); ebitda = r.get("ebitda")
    numf = lambda x: x if isinstance(x, (int, float)) else None
    equity = numf(equity); assets = numf(assets); liab = numf(liab)
    revenue = numf(revenue); profit = numf(profit); ebitda = numf(ebitda)
    # Capital próprio: no balancete a classe 5 NÃO inclui o resultado do período → somar; na IES já inclui.
    if dt == "balancete" and equity is not None and profit is not None:
        net_worth = round(equity + profit, 2)
    elif equity is not None:
        net_worth = equity
    elif assets is not None and liab is not None:
        net_worth = round(assets - liab, 2)
    else:
        net_worth = None
    if net_worth is None and profit is None and revenue is None:
        return None
    return {"year": top_year, "doc_type": dt, "doc_label": DOC_LABEL.get(dt, "Documento"),
            "assets": assets, "liabilities": liab, "equity": equity, "net_worth": net_worth,
            "annual_revenue": revenue, "annual_profit": profit, "ebitda": ebitda,
            "reconciled": r.get("reconciled")}


async def get_erp_financial_context(user_id: str, cid: Optional[str]):
    if not cid:
        return None
    ctx = await db.erp_financial_contexts.find_one(
        {"user_id": user_id, "company_id": cid, "active": True},
        {"_id": 0},
    )
    return ctx or None


def merge_financial_profile(profile: Optional[dict], erp_ctx: Optional[dict]):
    merged = dict(profile or {})
    if not erp_ctx or not erp_ctx.get("active"):
        return merged, None
    for key in ("monthly_revenue", "cash_balance", "variable_costs_pct", "total_debt"):
        value = erp_ctx.get(key)
        if isinstance(value, (int, float)):
            merged[key] = float(value)
    for key in ("fixed_costs", "assets", "liabilities"):
        value = erp_ctx.get(key)
        if isinstance(value, list) and value:
            merged[key] = value
    label = erp_ctx.get("source_label") or f"Sistema de gestão · {erp_ctx.get('system_name') or 'ERP'}"
    merged["_context_source_label"] = label
    merged["_context_updated_at"] = erp_ctx.get("updated_at")
    return merged, label

def compute_valuation_annual(fin: dict, cash: float = 0.0):
    """Valuation com base em figuras ANUAIS reais de documentos oficiais: base patrimonial + goodwill de rendimento."""
    net_worth = fin.get("net_worth") or 0.0
    annual_profit = fin.get("annual_profit") or 0.0
    annual_revenue = fin.get("annual_revenue") or 0.0
    margin = (annual_profit / annual_revenue * 100.0) if annual_revenue > 0 else 0.0
    mult = 0.0
    if annual_profit > 0:
        mult = 2.0
        if margin >= 10: mult += 0.5
        if margin >= 20: mult += 0.5
        if margin >= 30: mult += 0.5
    goodwill = max(0.0, annual_profit) * mult
    floor = net_worth if net_worth > 0 else 0.0
    value = round(max(floor + goodwill, cash or 0.0), 2)
    src = f"{fin['doc_label']} {fin['year']}"
    method = (f"patrimonial + rendimento (com base na tua {src})" if goodwill > 0
              else f"patrimonial (com base na tua {src})")
    return {"value": value, "net_worth": round(net_worth, 2), "annual_profit": round(annual_profit, 2),
            "multiple": mult, "goodwill": round(goodwill, 2), "method": method}


def compute_value_generic(net_worth, annual_profit, annual_revenue, cash=0.0):
    """Núcleo do valuation, agnóstico à fonte (manual ou documento): base patrimonial + goodwill de rendimento."""
    net_worth = net_worth or 0.0; annual_profit = annual_profit or 0.0; annual_revenue = annual_revenue or 0.0
    margin = (annual_profit / annual_revenue * 100.0) if annual_revenue > 0 else 0.0
    mult = 0.0
    if annual_profit > 0:
        mult = 2.0
        if margin >= 10: mult += 0.5
        if margin >= 20: mult += 0.5
        if margin >= 30: mult += 0.5
    goodwill = max(0.0, annual_profit) * mult
    floor = net_worth if net_worth > 0 else 0.0
    value = round(max(floor + goodwill, cash or 0.0), 2)
    return {"value": value, "net_worth": round(net_worth, 2), "annual_profit": round(annual_profit, 2),
            "multiple": mult, "goodwill": round(goodwill, 2)}


def value_multiple(margin_pct):
    """Múltiplo setorial-agnóstico usado por TODO o valuation (fonte única)."""
    mult = 2.0
    if margin_pct >= 10: mult += 0.5
    if margin_pct >= 20: mult += 0.5
    if margin_pct >= 30: mult += 0.5
    return mult


def required_performance_for_value(target_value, net_worth, assumed_margin_pct, cash=0.0):
    """Engenharia inversa: dado o VALOR-alvo da empresa e uma margem líquida assumida,
    devolve o lucro/faturação anuais necessários usando o MESMO motor de avaliação
    (valor = base patrimonial + lucro anual x múltiplo). Não é regra de três: o múltiplo
    depende da margem, logo faturação/margem/rentabilidade entram no cálculo."""
    floor = net_worth if net_worth and net_worth > 0 else 0.0
    mult = value_multiple(assumed_margin_pct)
    goodwill_needed = target_value - floor
    if goodwill_needed <= 0:
        return {"reached": True, "assumed_margin": round(assumed_margin_pct, 1), "multiple": mult,
                "required_profit": 0.0, "required_revenue": 0.0, "required_monthly_revenue": 0.0,
                "goodwill_needed": 0.0}
    req_profit = goodwill_needed / mult
    req_revenue = (req_profit / (assumed_margin_pct / 100.0)) if assumed_margin_pct and assumed_margin_pct > 0 else None
    return {"reached": False, "assumed_margin": round(assumed_margin_pct, 1), "multiple": mult,
            "required_profit": round(req_profit, 2),
            "required_revenue": round(req_revenue, 2) if req_revenue is not None else None,
            "required_monthly_revenue": round(req_revenue / 12, 2) if req_revenue is not None else None,
            "goodwill_needed": round(goodwill_needed, 2)}


# Múltiplos setoriais sugeridos (motor de avaliação híbrido). Faixas conservadoras para PME.
SECTOR_MULTIPLES = {
    "restaur":   {"label": "Restauração / Alimentação", "revenue": 0.8, "ebitda": 3.5},
    "aliment":   {"label": "Alimentação", "revenue": 0.8, "ebitda": 3.5},
    "café":      {"label": "Restauração / Cafetaria", "revenue": 0.8, "ebitda": 3.0},
    "constru":   {"label": "Construção", "revenue": 0.6, "ebitda": 3.5},
    "imobil":    {"label": "Imobiliário", "revenue": 1.5, "ebitda": 6.0},
    "software":  {"label": "Software / SaaS", "revenue": 2.5, "ebitda": 8.0},
    "tecnolog":  {"label": "Tecnologia", "revenue": 2.2, "ebitda": 7.5},
    "saas":      {"label": "SaaS", "revenue": 3.0, "ebitda": 8.0},
    "ecommerce": {"label": "E-commerce", "revenue": 1.2, "ebitda": 6.0},
    "comércio":  {"label": "Comércio / Retalho", "revenue": 0.7, "ebitda": 4.0},
    "retalho":   {"label": "Retalho", "revenue": 0.7, "ebitda": 4.0},
    "loja":      {"label": "Comércio / Loja", "revenue": 0.7, "ebitda": 4.0},
    "consultor": {"label": "Consultoria / Serviços", "revenue": 1.2, "ebitda": 5.0},
    "serviç":    {"label": "Serviços", "revenue": 1.2, "ebitda": 5.0},
    "saúde":     {"label": "Saúde / Clínica", "revenue": 1.4, "ebitda": 6.0},
    "clínic":    {"label": "Saúde / Clínica", "revenue": 1.4, "ebitda": 6.0},
    "indústr":   {"label": "Indústria", "revenue": 1.0, "ebitda": 5.0},
    "transport": {"label": "Transportes / Logística", "revenue": 0.9, "ebitda": 4.5},
    "turismo":   {"label": "Turismo / Hotelaria", "revenue": 1.5, "ebitda": 6.0},
}
_DEFAULT_MULT = {"label": "Geral", "revenue": 1.0, "ebitda": 5.0}


def suggest_multiples(sector: str, currency: str = "EUR"):
    """Sugere múltiplos de Faturação e de EBITDA com base no setor e região.
    O utilizador pode sempre ajustar manualmente."""
    s = (sector or "").lower()
    matched = _DEFAULT_MULT
    for key, val in SECTOR_MULTIPLES.items():
        if key in s:
            matched = val
            break
    region_factor = 0.9 if currency == "BRL" else 1.0  # ajuste de mercado emergente
    region = "Brasil" if currency == "BRL" else "Portugal / Europa"
    rev = round(matched["revenue"] * region_factor, 2)
    ebd = round(matched["ebitda"] * region_factor, 2)
    return {
        "sector_label": matched["label"], "region": region,
        "revenue": {"suggested": rev, "min": round(max(0.3, rev * 0.5), 2), "max": round(rev * 1.8, 2)},
        "ebitda": {"suggested": ebd, "min": round(max(2.0, ebd * 0.6), 2), "max": round(ebd * 1.6, 2)},
    }



async def build_snapshot(user_id: str):
    company = await resolve_company(user_id) or {}
    cid = str(company["_id"]) if company.get("_id") else None
    entries = await db.entries.find({"user_id": user_id, "company_id": cid}, {"type": 1, "amount": 1, "date": 1, "category": 1}).to_list(5000) if cid else []
    now = datetime.now(timezone.utc)
    month_key = now.strftime("%Y-%m")
    income = sum(e["amount"] for e in entries if e["type"] == "income")
    expense = sum(e["amount"] for e in entries if e["type"] == "expense")
    m_income = sum(e["amount"] for e in entries if e["type"] == "income" and str(e.get("date", "")).startswith(month_key))
    m_expense = sum(e["amount"] for e in entries if e["type"] == "expense" and str(e.get("date", "")).startswith(month_key))
    net = income - expense
    m_net = m_income - m_expense
    base_profile = await db.financial_profiles.find_one({"user_id": user_id, "company_id": cid}) if cid else None
    erp_ctx = await get_erp_financial_context(user_id, cid)
    profile, active_profile_label = merge_financial_profile(base_profile, erp_ctx)
    bal = compute_balance(company, profile, net)
    bank = bal["cash"]
    monthly_burn = m_expense if m_expense > 0 else (expense / 12 if expense else 1)
    runway = bank / monthly_burn if monthly_burn > 0 else 99
    profit_margin = (m_net / m_income * 100) if m_income > 0 else 0
    tax_reserve = float(company.get("monthly_tax_estimate", 0))
    payroll = sum(e["amount"] for e in entries if e["type"] == "expense" and "salári" in str(e.get("category", "")).lower())
    currency = company.get("currency", "EUR")

    vitals = [
        {"key": "cashflow", "label": "Fluxo de Caixa", "value": round(m_net, 2), "unit": CURRENCY_SYMBOL.get(currency, "€"),
         "status": rag(m_net, 0.01, -monthly_burn * 0.5), "hint": "Entradas menos saídas este mês"},
        {"key": "profit", "label": "Lucro", "value": round(profit_margin, 1), "unit": "%",
         "status": rag(profit_margin, 10, 3), "hint": "Margem de lucro mensal"},
        {"key": "clients", "label": "Clientes", "value": int(company.get("clients_count", 0)), "unit": "",
         "status": rag(company.get("clients_count", 0), 10, 3), "hint": "Base de clientes ativa"},
        {"key": "tax", "label": "Impostos", "value": round(tax_reserve, 2), "unit": CURRENCY_SYMBOL.get(currency, "€"),
         "status": rag(bank - tax_reserve, 0.01, -1), "hint": "Reserva vs estimativa fiscal"},
        {"key": "employees", "label": "Funcionários", "value": int(company.get("employees_count", 0)), "unit": "",
         "status": rag(m_income - payroll, 0.01, -1) if payroll else "green", "hint": "Custo de equipa sustentável"},
        {"key": "bank", "label": "Banco", "value": round(bank, 2), "unit": CURRENCY_SYMBOL.get(currency, "€"),
         "status": rag(bank, monthly_burn * 3, 0), "hint": "Saldo bancário estimado"},
        {"key": "risk", "label": "Risco", "value": round(runway, 1), "unit": "meses",
         "status": rag(runway, 6, 3), "hint": "Meses de autonomia de caixa"},
    ]
    status_score = {"green": 100, "amber": 55, "red": 20}
    health = round(sum(status_score[v["status"]] for v in vitals) / len(vitals))
    annual_profit = net if net > 0 else 0
    val = compute_valuation(profile, bal)
    company_value = val["value"]
    fin = await latest_official_financials(user_id, cid)
    has_doc = bool(fin and any(fin.get(k) is not None for k in ("net_worth", "annual_profit", "annual_revenue")))
    has_manual = bool(profile)
    financials_source = None; has_official = has_doc; value_sources = None
    if has_doc or has_manual:
        doc_label = f"{fin['doc_label']} {fin['year']}" if fin else None
        man_nw = bal["net_worth"] if has_manual else None
        man_profit = val["annual_profit"] if has_manual else None
        man_rev = (float(profile.get("monthly_revenue", 0) or 0) * 12) if has_manual else None
        manual_label = active_profile_label or "os teus dados (Perfil Financeiro)"
        def _pick(dv, mv):
            if has_doc and isinstance(dv, (int, float)):
                return dv, doc_label
            if isinstance(mv, (int, float)):
                return mv, manual_label
            return None, None
        nw, s_nw = _pick(fin.get("net_worth") if fin else None, man_nw)
        profit, s_profit = _pick(fin.get("annual_profit") if fin else None, man_profit)
        rev, s_rev = _pick(fin.get("annual_revenue") if fin else None, man_rev)
        ta, s_ta = _pick(fin.get("assets") if fin else None, bal["total_assets"] if has_manual else None)
        tl, s_tl = _pick(fin.get("liabilities") if fin else None, bal["total_liabilities"] if has_manual else None)
        cash = bal["cash"]
        nw_final = nw if nw is not None else bal["net_worth"]
        bal = {"cash": cash,
               "total_assets": round(ta, 2) if ta is not None else bal["total_assets"],
               "total_liabilities": round(tl, 2) if tl is not None else bal["total_liabilities"],
               "net_worth": round(nw_final, 2)}
        g = compute_value_generic(nw_final, profit, rev, cash)
        company_value = g["value"]
        method = "patrimonial + rendimento" if g["goodwill"] > 0 else "patrimonial"
        if has_doc:
            method += f" (com base na tua {doc_label})"
        val = {"value": g["value"], "net_worth": g["net_worth"], "annual_profit": g["annual_profit"],
               "annual_revenue": round(rev, 2) if isinstance(rev, (int, float)) else None,
               "multiple": g["multiple"], "goodwill": g["goodwill"], "method": method}
        financials_source = doc_label
        value_sources = {"patrimonio": s_nw, "lucro": s_profit, "faturacao": s_rev, "ativos": s_ta, "passivos": s_tl}
    has_balance_final = has_doc or has_manual
    dna = await db.ceo_dna.find_one({"user_id": user_id}) or {}
    goal_value = float(dna.get("target_revenue", 0)) or 1000000
    progress = min(100, round(company_value / goal_value * 100)) if goal_value else 0
    equity_progress = min(100, round(bal["net_worth"] / goal_value * 100)) if goal_value and bal["net_worth"] > 0 else 0

    return {
        "health": health, "vitals": vitals, "currency": currency,
        "currency_symbol": CURRENCY_SYMBOL.get(currency, "€"),
        "company_name": company.get("name", "A minha empresa"),
        "company_value": company_value, "goal_value": goal_value, "progress": progress,
        "valuation": val,
        "cash_balance": round(bank, 2), "monthly_net": round(m_net, 2),
        "monthly_income": round(m_income, 2), "monthly_expense": round(m_expense, 2),
        "runway": round(runway, 1), "profit_margin": round(profit_margin, 1),
        "total_income": round(income, 2), "total_expense": round(expense, 2),
        "cash_available": bal["cash"], "total_assets": bal["total_assets"],
        "total_liabilities": bal["total_liabilities"], "net_worth": bal["net_worth"],
        "has_balance": has_balance_final, "equity_progress": equity_progress,
        "financials_source": financials_source, "has_official": has_official,
        "value_sources": value_sources,
        "financial_context_source": active_profile_label,
    }

async def record_equity(user_id: str, cid, snap: dict):
    """Record this month's net worth so the panel can show equity evolution."""
    if not cid or not snap.get("has_balance"):
        return
    month = datetime.now(timezone.utc).strftime("%Y-%m")
    await db.equity_history.update_one(
        {"user_id": user_id, "company_id": cid, "month": month},
        {"$set": {"user_id": user_id, "company_id": cid, "month": month,
                  "net_worth": round(snap.get("net_worth", 0), 2),
                  "total_assets": round(snap.get("total_assets", 0), 2),
                  "total_liabilities": round(snap.get("total_liabilities", 0), 2),
                  "company_value": round(snap.get("company_value", 0), 2),
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True)


async def get_equity_history(user_id: str, cid):
    if not cid:
        return []
    rows = await db.equity_history.find({"user_id": user_id, "company_id": cid}).sort("month", 1).to_list(24)
    return rows[-12:]


MODE_PROMPTS = {
    "conservador": "És prudente e avesso ao risco. Priorizas estabilidade, reservas de caixa e evitas dívida.",
    "crescimento": "És focado em crescimento sustentável. Equilibras oportunidade e risco.",
    "agressivo": "És ambicioso e orientado a resultados rápidos. Aceitas mais risco por retorno maior.",
    "familiar": "És equilibrado, valorizas qualidade de vida, tempo com a família e sustentabilidade do negócio.",
    "startup": "És orientado a escala, produto e captação. Pensas em métricas de crescimento e runway.",
    "investidor": "Pensas como investidor: retorno sobre capital, valor da empresa e saída (exit).",
}

PROFILE_LABELS = {
    "activity": "O que a empresa faz",
    "years_active": "Anos de atividade",
    "location": "Localização",
    "business_model": "Como ganha dinheiro",
    "avg_price": "Preço médio do produto/serviço",
    "biggest_client_pct": "Peso do maior cliente nas vendas (%)",
    "client_recurrence": "Os clientes voltam a comprar",
    "founder_dependency": "A empresa funciona sem o dono",
    "debt": "Dívidas / empréstimos",
    "biggest_cost": "Maior custo mensal",
    "supplier_dependency": "Depende muito de um fornecedor",
    "seasonality": "Meses fortes ou fracos",
    "cae": "CAE (código de atividade)",
    "main_goal": "Objetivo com a empresa",
    "personal_goal": "Objetivo pessoal do dono",
    "advantage": "O que a distingue da concorrência",
    "main_worry": "Maior preocupação atual",
}

async def build_system_prompt(user_id: str, user_name: str):
    settings = await db.settings.find_one({"user_id": user_id}) or {}
    mode = settings.get("ceo_mode", "crescimento")
    tone = settings.get("briefing_tone", "direto")
    dna = await db.ceo_dna.find_one({"user_id": user_id}) or {}
    memories = await db.memories.find({"user_id": user_id}).to_list(100)
    snap = await build_snapshot(user_id)
    company = await resolve_company(user_id)
    cid = str(company["_id"]) if company and company.get("_id") else None
    prof = (company or {}).get("profile", {}) or {}
    sector = (company or {}).get("sector") or prof.get("activity") or ""
    cae = prof.get("cae")
    sector_line = (f"Esta empresa opera no setor: {sector}" + (f" (CAE {cae})" if cae else "") + ".") if sector else "O setor da empresa ainda NÃO está indicado."
    prof_txt = "\n".join(f"{lbl}: {prof.get(k)}" for k, lbl in PROFILE_LABELS.items() if prof.get(k) not in (None, "", 0)) or "(o empresário ainda não preencheu o perfil da empresa)"
    mem_txt = "\n".join(f"- {m['content']}" for m in memories) or "- (ainda sem memórias registadas)"
    vitals_txt = "\n".join(f"- {v['label']}: {v['value']}{v['unit']} [{v['status']}]" for v in snap["vitals"])
    import json as _json
    _docs = await db.documents.find({"user_id": user_id, "is_deleted": False}).sort("created_at", -1).to_list(12)
    _dlines, _figs = [], {}
    for _d in _docs:
        _a = _d.get("analysis") or {}
        if _a.get("summary"):
            _dlines.append(f"- {_d.get('original_filename', 'documento')} [{_d.get('doc_type', 'outro')}]: {_a.get('summary')}")
        for _k, _v in (_a.get("figures") or {}).items():
            if isinstance(_v, (int, float)) and _v and _k not in _figs:
                _figs[_k] = _v
    docs_block = ("\n".join(_dlines) + (("\nNúmeros extraídos dos documentos: " + _json.dumps(_figs, ensure_ascii=False)) if _figs else "")) if _dlines else "(o empresário ainda não carregou relatórios ou documentos)"
    erp_ctx = await get_erp_financial_context(user_id, cid)
    if erp_ctx:
        erp_fixed = ", ".join(f"{c.get('name')}: {c.get('amount')}€" for c in (erp_ctx.get("fixed_costs") or [])[:6]) or "sem custos fixos detalhados"
        works_info = erp_ctx.get("works_summary") or {}
        budgets_info = erp_ctx.get("budgets_summary") or {}
        exp_info = erp_ctx.get("expenses_breakdown") or {}
        exp_txt = ", ".join(f"{k}: {v}€" for k, v in exp_info.items()) or "sem despesas discriminadas"
        
        erp_block = (
            f"Sistema: {erp_ctx.get('system_name') or 'Obelisco Manager 360'}\n"
            f"Fonte ativa: {erp_ctx.get('source_label') or 'Obelisco Manager · Nuvem 360°'}\n"
            f"Última Sincronização: {erp_ctx.get('updated_at') or 'n/d'}\n"
            f"• TESOURARIA & SALDO REAL: {erp_ctx.get('cash_balance', 'n/d')}€\n"
            f"• FATURAÇÃO DO MÊS: {erp_ctx.get('monthly_revenue', 'n/d')}€ | EMITIDO NO ANO: {erp_ctx.get('annual_emitted_revenue', 'n/d')}€ (Meta 350.000€: {erp_ctx.get('annual_goal_progress_pct', 'n/d')}% concluído)\n"
            f"• CONTAS A RECEBER: {erp_ctx.get('amount_to_receive', 0)}€ (Vencido em atraso: {erp_ctx.get('overdue_to_receive', 0)}€)\n"
            f"• CONTAS A PAGAR: {erp_ctx.get('amount_to_pay', 0)}€ | DÍVIDA TOTAL: {erp_ctx.get('total_debt', 0)}€\n"
            f"• CUSTOS FIXOS TOTAIS: {erp_ctx.get('total_fixed_costs', 0)}€/mês ({erp_fixed})\n"
            f"• DESPESAS OPERACIONAIS DO MÊS: {exp_txt}\n"
            f"• EQUIPA & FUNCIONÁRIOS: {erp_ctx.get('active_employees_count', 0)} colaboradores ativos (Custo Mensal da Folha Salarial: {erp_ctx.get('payroll_monthly_cost', 0)}€)\n"
            f"• OBRAS & PROJETOS: {works_info.get('total_count', 0)} obras totais ({works_info.get('active_count', 0)} em curso, lucro estimado acumulado: {works_info.get('estimated_profit', 0)}€)\n"
            f"• COMERCIAL & ORÇAMENTOS: {budgets_info.get('total_budgets', 0)} orçamentos emitidos ({budgets_info.get('total_proposals', 0)} propostas)"
        )
    else:
        erp_block = "(sem integração ativa com sistema de gestão)"
    return (
        f"És o CEO AI 2.0 — o Diretor Executivo Digital de {user_name}. NÃO és um chatbot nem um assistente técnico: "
        f"és um CEO experiente que já geriu centenas de empresas e que agora toma decisões LADO A LADO com este empresário. "
        f"A tua personalidade é experiente, calma, objectiva e confiante. {MODE_PROMPTS.get(mode, MODE_PROMPTS['crescimento'])} Tom: {tone}.\n\n"
        f"### COMO RESPONDES (obrigatório)\n"
        f"NUNCA respondas apenas com teoria e NUNCA digas 'depende'. Respondes sempre como um consultor executivo de topo, "
        f"tomando posição. Estrutura natural de cada resposta (sem cabeçalhos rígidos, de forma fluida e humana):\n"
        f"1) O QUE EU FARIA — a decisão concreta, na primeira pessoa e directa.\n"
        f"2) PORQUÊ — o raciocínio ligado aos números reais e aos objectivos pessoais do empresário.\n"
        f"3) RISCOS — o que pode correr mal.\n"
        f"4) ALTERNATIVAS — 1 ou 2 caminhos possíveis.\n"
        f"Foca-te no FUTURO e nas decisões, não no passado. Sê conciso, calmo e confiante. Fala português europeu.\n\n"
        f"### ESPECIALIZAÇÃO NO SETOR (OBRIGATÓRIO — nunca generalizes)\n"
        f"{sector_line}\n"
        f"Age como um CEO que conhece PROFUNDAMENTE este setor específico. Todos os conselhos, "
        f"referências (margens típicas, ticket médio, custos-chave, sazonalidade), riscos, KPIs e boas práticas "
        f"DEVEM ser próprios deste setor — usa o vocabulário e a realidade de quem gere este tipo de negócio "
        f"(ex.: uma construtora fala de obras, adjudicações, mão-de-obra e materiais; um restaurante fala de "
        f"food cost, rotação de mesas, ementa e turnos). Compara sempre com as referências típicas DESTE setor e "
        f"evita conselhos genéricos que serviriam para qualquer empresa. Se o setor não estiver indicado, "
        f"recomenda ao empresário preenchê-lo na área Empresa.\n\n"
        f"### PERFIL (CEO DNA)\n"
        f"Sonho: {dna.get('dream', 'n/d')}\nFaturação desejada: {dna.get('target_revenue', 'n/d')}\n"
        f"Horas de trabalho: {dna.get('work_hours', 'n/d')}\nPlano de saída: {dna.get('exit_plan', 'n/d')}\n"
        f"Visão a 5 anos: {dna.get('five_year_vision', 'n/d')}\n\n"
        f"### MEMÓRIA (lembra-te disto sempre)\n{mem_txt}\n\n"
        f"### PERFIL DA EMPRESA (informação dada pelo empresário — usa-a sempre na tua análise)\n{prof_txt}\n\n"
        f"### DADOS ATIVOS DO SISTEMA DE GESTÃO (usar como contexto financeiro atual desta empresa)\n{erp_block}\n\n"
        f"### RELATÓRIOS E DOCUMENTOS CARREGADOS PELO EMPRESÁRIO (lê e usa estes dados reais; cita-os quando relevante)\n{docs_block}\n\n"
        f"### ESTADO ATUAL DA EMPRESA ({snap['company_name']})\n"
        f"Saúde: {snap['health']}/100\nCaixa: {snap['currency_symbol']}{snap['cash_balance']}\n"
        f"Resultado mensal: {snap['currency_symbol']}{snap['monthly_net']}\nAutonomia: {snap['runway']} meses\n"
        f"Valor da empresa: {snap['currency_symbol']}{snap['company_value']} (objetivo {snap['currency_symbol']}{snap['goal_value']}, {snap['progress']}%)\n"
        f"Sinais vitais:\n{vitals_txt}"
    )

MODELS_CASCADE = ["gemini-3.5-flash", "gemini-3.5-flash-lite", "gemini-flash-latest", "gemini-3-flash-preview", "gemini-3.6-flash", "gemini-3.7-flash"]
DEFAULT_LLM_MODEL = LITELLM_MODEL if LITELLM_API_KEY else "gemini-3.5-flash"
FINANCE_MODEL = LITELLM_MODEL if LITELLM_API_KEY else "gemini-3.5-flash"

PT_PT_INSTRUCTION = (
    "\n\n[DIRETIVA DE LINGUAGEM E REVISÃO OBRIGATÓRIA - PORTUGUÊS DE PORTUGAL]\n"
    "1. Todos os textos gerados (conteúdos de marketing, artigos do site, páginas, posts de redes sociais, emails, briefings, relatórios) "
    "devem ser estritamente em Português de Portugal (Português Europeu - PT-PT).\n"
    "2. Proibido usar gerúndios típicos do português do Brasil (usa 'estamos a fazer', 'a crescer', 'a planear' em vez de 'fazendo', 'crescendo', 'planejando').\n"
    "3. Usa vocabulário europeu correto: 'equipa' (nunca 'equipe'), 'contacto' (nunca 'contato'), 'connosco' (nunca 'conosco'), 'utilizador' (nunca 'usuário'), "
    "'publicar/partilhar' (nunca 'postar'), 'gerir' (nunca 'gerenciar'), 'planear' (nunca 'planejar'), 'ecrã' (nunca 'tela'), 'telemóvel' (nunca 'celular'), "
    "'registo/registar' (nunca 'cadastro/cadastrar'), 'facturação/faturação', 'otimização/optimização'.\n"
    "4. Revisão e Ortografia: Realiza uma revisão minuciosa para garantir acentuação perfeita e ZERO erros ortográficos ou gramaticais."
)

def sanitize_pt_pt(data):
    """Garante limpeza e correção determinística de termos para Português de Portugal."""
    if isinstance(data, str):
        text = data
        replacements = [
            (r"\bequipe\b", "equipa"),
            (r"\bEquipe\b", "Equipa"),
            (r"\bequipes\b", "equipas"),
            (r"\bEquipes\b", "Equipas"),
            (r"\bcontato\b", "contacto"),
            (r"\bContato\b", "Contacto"),
            (r"\bcontatos\b", "contactos"),
            (r"\bContatos\b", "Contactos"),
            (r"\bconosco\b", "connosco"),
            (r"\bConosco\b", "Connosco"),
            (r"\busuário\b", "utilizador"),
            (r"\bUsuário\b", "Utilizador"),
            (r"\busuários\b", "utilizadores"),
            (r"\bUsuários\b", "Utilizadores"),
            (r"\bgerenciar\b", "gerir"),
            (r"\bGerenciar\b", "Gerir"),
            (r"\bplanejar\b", "planear"),
            (r"\bPlanejar\b", "Planear"),
            (r"\bplanejamento\b", "planeamento"),
            (r"\bPlanejamento\b", "Planeamento"),
            (r"\bpostar\b", "publicar"),
            (r"\bPostar\b", "Publicar"),
            (r"\bcadastre-se\b", "registe-se"),
            (r"\bCadastre-se\b", "Registe-se"),
            (r"\bcadastro\b", "registo"),
            (r"\bCadastro\b", "Registo"),
            (r"\bcelular\b", "telemóvel"),
            (r"\bCelular\b", "Telemóvel"),
        ]
        import re
        for pat, repl in replacements:
            text = re.sub(pat, repl, text)
        return text
    elif isinstance(data, list):
        return [sanitize_pt_pt(item) for item in data]
    elif isinstance(data, dict):
        return {k: sanitize_pt_pt(v) for k, v in data.items()}
    return data

async def call_litellm_text(system: str, prompt_or_contents, model: str = None) -> str:
    """Chama o gateway LiteLLM através da interface compatível com OpenAI."""
    if not LITELLM_API_KEY:
        return ""
    target_model = model or LITELLM_MODEL or "Gemini Auto Key"
    full_system = (system or "") + PT_PT_INSTRUCTION
    
    if isinstance(prompt_or_contents, list):
        user_text = " ".join([str(p) for p in prompt_or_contents if isinstance(p, str)])
    else:
        user_text = str(prompt_or_contents)
        
    messages = [
        {"role": "system", "content": full_system},
        {"role": "user", "content": user_text}
    ]
    
    headers = {
        "Authorization": f"Bearer {LITELLM_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": target_model,
        "messages": messages,
        "temperature": 0.7,
    }
    endpoints = [
        f"{LITELLM_BASE_URL}/chat/completions",
        f"{LITELLM_BASE_URL}/v1/chat/completions",
    ]
    async with httpx.AsyncClient(timeout=45.0, verify=False) as client:
        for ep in endpoints:
            try:
                res = await client.post(ep, json=payload, headers=headers)
                if res.status_code == 200:
                    data = res.json()
                    choices = data.get("choices", [])
                    if choices:
                        content = choices[0].get("message", {}).get("content", "")
                        if content:
                            return sanitize_pt_pt(content.strip())
                else:
                    logger.warning(f"LiteLLM {ep} returned {res.status_code}: {res.text[:150]}")
            except Exception as e:
                logger.warning(f"LiteLLM call error on {ep}: {e}")
                continue
    return ""

async def stream_litellm_chat(system_instruction: str, contents, model: str = None):
    """Transmite a resposta do LiteLLM via SSE (OpenAI streaming)."""
    if not LITELLM_API_KEY:
        return
    target_model = model or LITELLM_MODEL or "Gemini Auto Key"
    full_system = (system_instruction or "") + PT_PT_INSTRUCTION
    
    if isinstance(contents, list):
        user_text = " ".join([str(p) for p in contents if isinstance(p, str)])
    else:
        user_text = str(contents)
        
    messages = [
        {"role": "system", "content": full_system},
        {"role": "user", "content": user_text}
    ]
    headers = {
        "Authorization": f"Bearer {LITELLM_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": target_model,
        "messages": messages,
        "temperature": 0.7,
        "stream": True,
    }
    endpoints = [
        f"{LITELLM_BASE_URL}/chat/completions",
        f"{LITELLM_BASE_URL}/v1/chat/completions",
    ]
    async with httpx.AsyncClient(timeout=60.0, verify=False) as client:
        for ep in endpoints:
            try:
                async with client.stream("POST", ep, json=payload, headers=headers) as response:
                    if response.status_code != 200:
                        continue
                    async for line in response.aiter_lines():
                        line = line.strip()
                        if not line:
                            continue
                        if line.startswith("data: "):
                            data_str = line[6:].strip()
                            if data_str == "[DONE]":
                                break
                            try:
                                chunk_json = json.loads(data_str)
                                choices = chunk_json.get("choices", [])
                                if choices:
                                    content = choices[0].get("delta", {}).get("content")
                                    if content:
                                        yield content
                            except Exception:
                                pass
                    return
            except Exception as e:
                logger.warning(f"LiteLLM stream error on {ep}: {e}")
                continue

async def ai_text(system: str, prompt_or_contents, model: str = DEFAULT_LLM_MODEL) -> str:
    # 1. Tentar primeiro o Gateway Central LiteLLM se configurado
    if LITELLM_API_KEY:
        try:
            lit_res = await call_litellm_text(system, prompt_or_contents, model=model)
            if lit_res:
                return lit_res
        except Exception as e:
            logger.warning(f"LiteLLM gateway call error: {e}")

    # 2. Fallback direto para Gemini se disponível
    if not genai_client:
        logger.error("Nenhum provedor de LLM configurado (LiteLLM / Gemini)")
        return ""
    full_system = (system or "") + PT_PT_INSTRUCTION
    contents = prompt_or_contents if isinstance(prompt_or_contents, list) else [prompt_or_contents]
    models_to_try = [model] + [m for m in MODELS_CASCADE if m != model]
    for m in models_to_try:
        try:
            res = await genai_client.aio.models.generate_content(
                model=m,
                contents=contents,
                config=types.GenerateContentConfig(
                    system_instruction=full_system,
                    temperature=0.7,
                )
            )
            text = (getattr(res, "text", "") or "").strip()
            if text:
                return sanitize_pt_pt(text)
        except Exception as e:
            logger.warning(f"ai_text model {m} failed: {e}")
            continue
    return ""

async def stream_gemini_chat(system_instruction: str, contents, model: str = DEFAULT_LLM_MODEL):
    # 1. Tentar primeiro o Gateway Central LiteLLM se configurado
    if LITELLM_API_KEY:
        yielded_any = False
        try:
            async for chunk in stream_litellm_chat(system_instruction, contents, model=model):
                yielded_any = True
                yield chunk
            if yielded_any:
                return
        except Exception as e:
            logger.warning(f"LiteLLM stream error: {e}")

    # 2. Fallback direto para Gemini se disponível
    if not genai_client:
        logger.error("Nenhum provedor de LLM configurado (LiteLLM / Gemini)")
        yield " [Provedor de LLM não configurado. Verifique as credenciais.]"
        return

    full_system = (system_instruction or "") + PT_PT_INSTRUCTION
    models_to_try = [model] + [m for m in MODELS_CASCADE if m != model]
    
    for m in models_to_try:
        try:
            response = await genai_client.aio.models.generate_content_stream(
                model=m,
                contents=contents,
                config=types.GenerateContentConfig(
                    system_instruction=full_system,
                    temperature=0.7,
                )
            )
            yielded_any = False
            async for chunk in response:
                if chunk.text:
                    yielded_any = True
                    yield chunk.text
            if yielded_any:
                return
        except Exception as e:
            logger.warning(f"stream_gemini_chat model {m} failed: {e}")
            continue

    yield " [Não foi possível gerar a resposta no momento. Tente novamente.]"

class GeminiChatWrapper:
    def __init__(self, system_instruction: str, model: str = DEFAULT_LLM_MODEL):
        self.system_instruction = system_instruction
        self.model = model

    def _build_parts(self, user_msg):
        text = user_msg if isinstance(user_msg, str) else getattr(user_msg, "text", str(user_msg))
        parts = []
        if hasattr(user_msg, "file_contents") and user_msg.file_contents:
            for fc in user_msg.file_contents:
                if hasattr(fc, "image_base64") and fc.image_base64:
                    parts.append(types.Part.from_bytes(data=base64.b64decode(fc.image_base64), mime_type="image/png"))
                elif hasattr(fc, "file_path") and fc.file_path:
                    try:
                        with open(fc.file_path, "rb") as f:
                            parts.append(types.Part.from_bytes(data=f.read(), mime_type=getattr(fc, "mime_type", "application/pdf")))
                    except Exception as e:
                        logger.error(f"file part error: {e}")
        parts.append(text)
        return parts

    async def send_message(self, user_msg) -> str:
        parts = self._build_parts(user_msg)
        return await ai_text(self.system_instruction, parts, model=self.model)

    async def stream_message(self, user_msg):
        parts = self._build_parts(user_msg)
        async for chunk in stream_gemini_chat(self.system_instruction, parts, model=self.model):
            yield TextDelta(chunk)

async def get_chat(user_id: str, user_name: str, session_id: str, vision: bool = False):
    sysmsg = await build_system_prompt(user_id, user_name)
    return GeminiChatWrapper(system_instruction=sysmsg, model=DEFAULT_LLM_MODEL)

async def make_briefing(user_id: str, user_name: str):
    settings = await db.settings.find_one({"user_id": user_id}) or {}
    count = settings.get("briefing_count", 4)
    snap = await build_snapshot(user_id)
    sysmsg = await build_system_prompt(user_id, user_name)
    hour = datetime.now(timezone.utc).hour
    greeting = "Bom dia" if hour < 12 else ("Boa tarde" if hour < 19 else "Boa noite")
    prompt = (
        f"Gera o briefing diário para o empresário. Devolve APENAS JSON válido no formato: "
        f'{{"greeting":"{greeting}, <nome>. ...","items":[{{"title":str,"detail":str,"priority":"alta"|"media"|"baixa","icon":"cash"|"profit"|"clients"|"tax"|"risk"|"opportunity"}}]}}. '
        f"Exatamente {count} itens, priorizados pelo que mais importa hoje. "
        f"O greeting deve ser uma frase humana e calorosa a começar com '{greeting}'. Detalhes curtos, orientados ao futuro e à ação. Sem texto fora do JSON."
    )
    data = await ai_json(sysmsg, prompt, model=DEFAULT_LLM_MODEL)
    if not data or not isinstance(data, dict):
        data = {"greeting": f"{greeting}, {user_name}. Aqui está o que precisa da sua atenção hoje.",
                "items": [{"title": "Ligue os seus dados", "detail": "Registe receitas e despesas para eu analisar a saúde da sua empresa.",
                           "priority": "alta", "icon": "opportunity"}]}
    data["health"] = snap.get("health", 0)
    return data

PRIORITY_COLOR = {"alta": "#EF4444", "media": "#F59E0B", "baixa": "#10B981"}

def build_briefing_html(name: str, data: dict, app_url: str):
    rows = ""
    for it in data.get("items", []):
        pc = PRIORITY_COLOR.get(it.get("priority", "media"), "#F59E0B")
        rows += f"""
        <tr><td style="padding:0 0 14px 0;">
          <table width="100%" cellpadding="0" cellspacing="0" style="background:#faf9f6;border:1px solid #eee;border-radius:12px;">
            <tr>
              <td width="6" style="background:{pc};border-radius:12px 0 0 12px;">&nbsp;</td>
              <td style="padding:14px 18px;">
                <div style="font-size:15px;font-weight:700;color:#18181b;">{it.get('title','')}</div>
                <div style="font-size:14px;color:#52525b;margin-top:4px;line-height:1.5;">{it.get('detail','')}</div>
              </td>
            </tr>
          </table>
        </td></tr>"""
    return f"""<!DOCTYPE html><html><body style="margin:0;background:#0b0c10;font-family:Arial,Helvetica,sans-serif;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#0b0c10;padding:32px 0;">
      <tr><td align="center">
        <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:18px;overflow:hidden;">
          <tr><td style="background:#0b0c10;padding:28px 32px;">
            <div style="color:#D4AF37;font-size:22px;font-weight:700;letter-spacing:1px;">CEO&nbsp;AI</div>
            <div style="color:#a1a1aa;font-size:11px;letter-spacing:2px;text-transform:uppercase;margin-top:2px;">Executivo Digital · Briefing Diário</div>
          </td></tr>
          <tr><td style="padding:32px;">
            <div style="font-size:22px;color:#18181b;font-weight:700;line-height:1.35;margin-bottom:8px;">{data.get('greeting','Bom dia')}</div>
            <div style="font-size:13px;color:#71717a;margin-bottom:22px;">Saúde da empresa: <strong style="color:#D4AF37;">{data.get('health',0)}/100</strong></div>
            <table width="100%" cellpadding="0" cellspacing="0">{rows}</table>
            <table width="100%" cellpadding="0" cellspacing="0" style="margin-top:22px;"><tr><td align="center">
              <a href="{app_url}" style="display:inline-block;background:#D4AF37;color:#0b0c10;text-decoration:none;font-weight:700;font-size:14px;padding:13px 28px;border-radius:999px;">Abrir o meu CEO AI 2.0</a>
            </td></tr></table>
          </td></tr>
          <tr><td style="padding:20px 32px;background:#faf9f6;border-top:1px solid #eee;">
            <div style="font-size:11px;color:#a1a1aa;">Recebes este email porque ativaste o briefing diário. Podes desativar em Personalização.</div>
          </td></tr>
        </table>
      </td></tr>
    </table></body></html>"""

async def send_email_raw(to_email: str, subject: str, html: str):
    smtp_host = os.environ.get("SMTP_HOST")
    if smtp_host:
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            
            smtp_port = int(os.environ.get("SMTP_PORT", 587))
            smtp_user = os.environ.get("SMTP_USER", "")
            smtp_pass = os.environ.get("SMTP_PASSWORD", "")
            from_email = os.environ.get("SMTP_FROM", smtp_user or "noreply@ceoai.local")
            
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = f"{EMAIL_FROM_NAME} <{from_email}>"
            msg["To"] = to_email
            msg.attach(MIMEText(html, "html"))
            
            def _send_smtp():
                with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
                    if os.environ.get("SMTP_USE_TLS", "true").lower() in ("true", "1"):
                        server.starttls()
                    if smtp_user and smtp_pass:
                        server.login(smtp_user, smtp_pass)
                    server.sendmail(from_email, [to_email], msg.as_string())
            
            await asyncio.to_thread(_send_smtp)
            logger.info(f"Email sent via SMTP to {to_email}: {subject}")
            return True
        except Exception as e:
            logger.error(f"SMTP send error: {e}")
            return False
            
    resend_key = os.environ.get("RESEND_API_KEY")
    if resend_key:
        try:
            from_email = os.environ.get("RESEND_FROM", "onboarding@resend.dev")
            payload = {
                "from": f"{EMAIL_FROM_NAME} <{from_email}>",
                "to": [to_email],
                "subject": subject,
                "html": html,
            }
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.post(
                    "https://api.resend.com/emails",
                    headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                    json=payload
                )
            resp.raise_for_status()
            logger.info(f"Email sent via Resend to {to_email}: {subject}")
            return True
        except Exception as e:
            logger.error(f"Resend send error: {e}")
            return False

    # Dev/Local Mode: Log email safely
    logger.info(f"[DEV EMAIL] To: {to_email} | Subject: {subject} | Body length: {len(html)} bytes")
    return True

# ---------------------------------------------------------------- password reset
def hash_reset_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()

async def create_password_reset(user_id: str) -> str:
    raw = secrets.token_urlsafe(32)
    await db.password_reset_tokens.insert_one({
        "user_id": user_id,
        "token_hash": hash_reset_token(raw),
        "used": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=1),
    })
    return raw

def build_reset_password_html(name: str, link: str) -> str:
    who = f", {name}" if name else ""
    return (f"<div style='font-family:Arial,sans-serif;max-width:560px;margin:auto;padding:24px'>"
            f"<h2 style='color:#0b0c10'>Redefinicao de senha - CEO AI 2.0</h2>"
            f"<p>Ola{who}, recebemos um pedido para redefinir a senha da tua conta CEO AI 2.0.</p>"
            f"<p style='margin:24px 0'><a href='{link}' style='display:inline-block;background:#3B82F6;color:#fff;text-decoration:none;font-weight:700;font-size:14px;padding:13px 28px;border-radius:999px;'>Definir nova senha</a></p>"
            f"<p style='font-size:13px;color:#71717a'>Esta ligacao e valida por 1 hora e so pode ser usada uma vez. Se nao pediste esta alteracao, ignora este email.</p>"
            f"<p style='font-size:12px;color:#a1a1aa;word-break:break-all'>{link}</p>"
            f"</div>")

async def send_password_reset_email(user_doc: dict) -> bool:
    raw = await create_password_reset(str(user_doc["_id"]))
    frontend = os.environ.get("FRONTEND_URL", "").rstrip("/")
    link = f"{frontend}/reset-password?token={raw}"
    html = build_reset_password_html(user_doc.get("name", ""), link)
    return await send_email_raw(user_doc.get("email", ""), "Redefinicao de senha - CEO AI 2.0", html)


async def send_daily_briefings():
    today = datetime.now(timezone.utc).date().isoformat()
    cursor = db.settings.find({"email_briefing": True})
    async for s in cursor:
        uid = s.get("user_id")
        if not uid:
            continue
        claim = await db.settings.update_one(
            {"user_id": uid, "email_briefing": True, "last_briefing_email_date": {"$ne": today}},
            {"$set": {"last_briefing_email_date": today}})
        if claim.modified_count != 1:
            continue
        try:
            u = await db.users.find_one({"_id": ObjectId(uid)})
            if not u:
                continue
            data = await make_briefing(uid, u.get("name", ""))
            html = build_briefing_html(u.get("name", ""), data, os.environ.get("FRONTEND_URL", ""))
            await send_email_raw(u["email"], "O teu briefing diário — CEO AI 2.0", html)
        except Exception as e:
            logger.error(f"daily briefing error for {uid}: {e}")


MONTH_ABBR = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]


async def compute_value_alert(user_id: str, cid):
    company = await resolve_company(user_id) or {}
    sym = CURRENCY_SYMBOL.get(company.get("currency", "EUR"), "€")
    if not cid:
        return {"has_alert": False, "currency_symbol": sym}
    rows = await db.equity_history.find({"user_id": user_id, "company_id": cid}).sort("month", 1).to_list(24)
    if len(rows) < 2:
        return {"has_alert": False, "currency_symbol": sym}
    cur, prev = rows[-1], rows[-2]
    cv, pv = cur.get("company_value"), prev.get("company_value")
    if cv is None or pv is None:
        return {"has_alert": False, "currency_symbol": sym}
    delta = round(cv - pv, 2)
    pct = round(delta / pv * 100, 1) if pv else None
    return {"has_alert": abs(delta) >= 1, "current": cv, "previous": pv, "delta": delta, "pct": pct,
            "direction": "up" if delta >= 0 else "down", "month": cur["month"],
            "month_label": MONTH_ABBR[int(cur["month"][5:7]) - 1],
            "prev_month_label": MONTH_ABBR[int(prev["month"][5:7]) - 1],
            "currency_symbol": sym}


def build_value_alert_html(name: str, alert: dict, app_url: str):
    sym = alert.get("currency_symbol", "€")
    up = alert["direction"] == "up"
    color = "#10B981" if up else "#EF4444"
    fnum = lambda v: f"{int(round(v)):,}".replace(",", " ")
    who = f", {name}" if name else ""
    pct = f" ({'+' if up else '−'}{abs(alert['pct'])}%)" if alert.get("pct") is not None else ""
    verb = "subiu" if up else "desceu"
    lead = ("Boas notícias" if up else "Atenção")
    note = ("O valor da tua empresa está a crescer. Continua a executar o plano."
            if up else "O valor da tua empresa desceu este mês. Vale a pena perceber porquê.")
    return f"""<!DOCTYPE html><html><body style="margin:0;background:#0b0c10;font-family:Arial,Helvetica,sans-serif;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#0b0c10;padding:32px 0;">
      <tr><td align="center">
        <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:18px;overflow:hidden;">
          <tr><td style="background:#0b0c10;padding:28px 32px;">
            <div style="color:#3B82F6;font-size:22px;font-weight:700;letter-spacing:1px;">CEO&nbsp;AI</div>
            <div style="color:#a1a1aa;font-size:11px;letter-spacing:2px;text-transform:uppercase;margin-top:2px;">Executivo Digital · Valor da Empresa</div>
          </td></tr>
          <tr><td style="padding:32px;">
            <div style="font-size:13px;color:#71717a;margin-bottom:6px;">{lead}{who}</div>
            <div style="font-size:26px;color:#18181b;font-weight:800;line-height:1.25;margin-bottom:6px;">
              A tua empresa vale {sym}{fnum(alert['current'])}
            </div>
            <div style="font-size:16px;color:{color};font-weight:700;margin-bottom:18px;">
              {verb} {sym}{fnum(abs(alert['delta']))}{pct} desde {alert.get('prev_month_label','')}
            </div>
            <div style="font-size:14px;color:#52525b;line-height:1.6;margin-bottom:24px;">{note}</div>
            <table width="100%" cellpadding="0" cellspacing="0"><tr><td align="center">
              <a href="{app_url}" style="display:inline-block;background:#3B82F6;color:#ffffff;text-decoration:none;font-weight:700;font-size:14px;padding:13px 28px;border-radius:999px;">Ver o valor da minha empresa</a>
            </td></tr></table>
          </td></tr>
          <tr><td style="padding:20px 32px;background:#faf9f6;border-top:1px solid #eee;">
            <div style="font-size:11px;color:#a1a1aa;">Recebes este resumo mensal do valor da tua empresa do CEO AI 2.0. Podes desativar em Personalização.</div>
          </td></tr>
        </table>
      </td></tr>
    </table></body></html>"""


async def send_monthly_value_alerts():
    pairs = await db.equity_history.aggregate([{"$group": {"_id": {"u": "$user_id", "c": "$company_id"}}}]).to_list(100000)
    for p in pairs:
        uid = p["_id"].get("u"); cid = p["_id"].get("c")
        if not uid or not cid:
            continue
        try:
            alert = await compute_value_alert(uid, cid)
            if not alert.get("has_alert"):
                continue
            s = await db.settings.find_one({"user_id": uid}) or {}
            if s.get("email_value_alert") is False:
                continue
            claim = await db.equity_history.update_one(
                {"user_id": uid, "company_id": cid, "month": alert["month"], "alert_emailed": {"$ne": True}},
                {"$set": {"alert_emailed": True}})
            if claim.modified_count != 1:
                continue
            u = await db.users.find_one({"_id": ObjectId(uid)})
            if not u or not u.get("email"):
                continue
            html = build_value_alert_html(u.get("name", ""), alert, os.environ.get("FRONTEND_URL", ""))
            subj = ("O valor da tua empresa subiu este mês — CEO AI 2.0" if alert["direction"] == "up"
                    else "O valor da tua empresa mudou este mês — CEO AI 2.0")
            await send_email_raw(u["email"], subj, html)
            try:
                await send_push_to_user(uid, subj, f"A tua empresa vale {alert['currency_symbol']}{int(round(alert['current']))}", "/valor")
            except Exception:
                pass
        except Exception as e:
            logger.error(f"monthly value alert error for {uid}: {e}")


def build_goal_alert_html(name: str, sym: str, current: float, target: float, pct: float, reached: bool, app_url: str):
    fnum = lambda v: f"{int(round(v)):,}".replace(",", " ")
    who = f", {name}" if name else ""
    color = "#10B981"
    lead = "Meta atingida!" if reached else "Estás quase lá"
    headline = ("A tua empresa atingiu a meta de valor" if reached
                else f"A tua empresa está a {int(round(pct))}% da tua meta de valor")
    note = ("Parabéns — o valor estimado da tua empresa alcançou a meta que definiste. "
            "Podes definir uma nova meta mais ambiciosa em Metas e Projeções."
            if reached else
            "Estás muito perto de atingir a meta de valor que definiste. Mantém o ritmo — o último empurrão conta.")
    return f"""<!DOCTYPE html><html><body style="margin:0;background:#0b0c10;font-family:Arial,Helvetica,sans-serif;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#0b0c10;padding:32px 0;">
      <tr><td align="center">
        <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:18px;overflow:hidden;">
          <tr><td style="background:#0b0c10;padding:28px 32px;">
            <div style="color:#3B82F6;font-size:22px;font-weight:700;letter-spacing:1px;">CEO&nbsp;AI&nbsp;2.0</div>
            <div style="color:#a1a1aa;font-size:11px;letter-spacing:2px;text-transform:uppercase;margin-top:2px;">Executivo Digital · Meta de Valor</div>
          </td></tr>
          <tr><td style="padding:32px;">
            <div style="font-size:13px;color:#71717a;margin-bottom:6px;">{lead}{who}</div>
            <div style="font-size:24px;color:#18181b;font-weight:800;line-height:1.3;margin-bottom:10px;">{headline}</div>
            <div style="font-size:16px;color:{color};font-weight:700;margin-bottom:6px;">
              Valor atual: {sym}{fnum(current)} · Meta: {sym}{fnum(target)}
            </div>
            <div style="height:10px;background:#eee;border-radius:999px;overflow:hidden;margin:14px 0 22px;">
              <div style="height:10px;width:{min(100, int(round(pct)))}%;background:{color};"></div>
            </div>
            <div style="font-size:14px;color:#52525b;line-height:1.6;margin-bottom:24px;">{note}</div>
            <table width="100%" cellpadding="0" cellspacing="0"><tr><td align="center">
              <a href="{app_url}/meta" style="display:inline-block;background:#3B82F6;color:#ffffff;text-decoration:none;font-weight:700;font-size:14px;padding:13px 28px;border-radius:999px;">Ver a minha projeção</a>
            </td></tr></table>
          </td></tr>
          <tr><td style="padding:20px 32px;background:#faf9f6;border-top:1px solid #eee;">
            <div style="font-size:11px;color:#a1a1aa;">Recebes este aviso quando o valor da tua empresa se aproxima ou atinge a meta. Podes desativar em Personalização.</div>
          </td></tr>
        </table>
      </td></tr>
    </table></body></html>"""


async def compute_goal_progress(user_id: str, cid):
    """Progresso da meta de valor: valor atual (motor) vs meta guardada."""
    g = await db.goals.find_one({"user_id": user_id, "company_id": cid}) or {}
    tv = float(g.get("target_value") or 0)
    if tv <= 0:
        return None
    snap = await build_snapshot(user_id)
    cv = float(snap.get("company_value", 0) or 0)
    pct = round(cv / tv * 100, 1) if tv else 0
    return {"goal": g, "current": cv, "target": tv, "pct": pct,
            "reached": cv >= tv, "currency_symbol": snap.get("currency_symbol", "€"),
            "company_name": snap.get("company_name", "")}


async def send_goal_alerts():
    """Cron: avisa por email quando o valor da empresa se aproxima (>=90%) ou atinge (100%) a meta.
    Idempotente via flags no documento da meta."""
    goals = await db.goals.find({"target_value": {"$gt": 0}}).to_list(100000)
    for g in goals:
        uid = g.get("user_id"); cid = g.get("company_id")
        if not uid:
            continue
        try:
            prog = await compute_goal_progress(uid, cid)
            if not prog:
                continue
            reached = prog["reached"]; pct = prog["pct"]
            flag = None
            if reached and not g.get("goal_reached_emailed"):
                flag = "goal_reached_emailed"
            elif not reached and pct >= 90 and not g.get("goal_near_emailed"):
                flag = "goal_near_emailed"
            if not flag:
                continue
            s = await db.settings.find_one({"user_id": uid}) or {}
            if s.get("email_value_alert") is False:
                continue
            claim = await db.goals.update_one(
                {"_id": g["_id"], flag: {"$ne": True}}, {"$set": {flag: True}})
            if claim.modified_count != 1:
                continue
            u = await db.users.find_one({"_id": ObjectId(uid)})
            if not u or not u.get("email"):
                continue
            html = build_goal_alert_html(u.get("name", ""), prog["currency_symbol"], prog["current"],
                                         prog["target"], pct, reached, os.environ.get("FRONTEND_URL", ""))
            subj = ("Atingiste a tua meta de valor — CEO AI 2.0" if reached
                    else "Estás quase a atingir a tua meta de valor — CEO AI 2.0")
            await send_email_raw(u["email"], subj, html)
            try:
                await send_push_to_user(uid, subj, f"Valor atual {prog['currency_symbol']}{int(round(prog['current']))} de {prog['currency_symbol']}{int(round(prog['target']))}", "/meta")
            except Exception:
                pass
        except Exception as e:
            logger.error(f"goal alert error for {uid}: {e}")



async def ensure_vapid():
    cfg = await db.app_config.find_one({"_id": "vapid"})
    if cfg and cfg.get("public") and cfg.get("private"):
        return cfg
    import base64
    from cryptography.hazmat.primitives.asymmetric import ec
    from cryptography.hazmat.primitives import serialization
    pk = ec.generate_private_key(ec.SECP256R1())
    priv = base64.urlsafe_b64encode(pk.private_numbers().private_value.to_bytes(32, "big")).decode().rstrip("=")
    pub = base64.urlsafe_b64encode(pk.public_key().public_bytes(serialization.Encoding.X962, serialization.PublicFormat.UncompressedPoint)).decode().rstrip("=")
    cfg = {"_id": "vapid", "public": pub, "private": priv, "subject": os.environ.get("VAPID_SUBJECT", "mailto:admin@ceo-ai.app")}
    await db.app_config.update_one({"_id": "vapid"}, {"$set": cfg}, upsert=True)
    return cfg


def _webpush_send(sub, payload, priv, subject):
    from pywebpush import webpush, WebPushException
    try:
        webpush(subscription_info=sub, data=payload, vapid_private_key=priv, vapid_claims={"sub": subject})
        return True
    except WebPushException as e:
        return getattr(getattr(e, "response", None), "status_code", None)


async def send_push_to_user(user_id: str, title: str, body: str, url: str = "/", actions=None, extra=None):
    import asyncio as _asyncio, json as _json
    cfg = await ensure_vapid()
    subs = await db.push_subscriptions.find({"user_id": user_id}).to_list(50)
    data = {"title": title, "body": body, "url": url}
    if actions:
        data["actions"] = actions
    if extra:
        data.update(extra)
    payload = _json.dumps(data)
    sent = 0
    for s in subs:
        res = await _asyncio.to_thread(_webpush_send, {"endpoint": s["endpoint"], "keys": s["keys"]}, payload, cfg["private"], cfg["subject"])
        if res is True:
            sent += 1
        elif res in (404, 410):
            await db.push_subscriptions.delete_one({"_id": s["_id"]})
    return sent

async def ai_json(system: str, prompt_or_contents, model: str = DEFAULT_LLM_MODEL):
    # 1. Tentar primeiro via LiteLLM se configurado
    if LITELLM_API_KEY:
        try:
            json_system = (system or "") + "\nResponde APENAS com um objeto JSON válido, sem texto extra nem markdown."
            raw_text = await call_litellm_text(json_system, prompt_or_contents, model=model)
            if raw_text:
                t = raw_text.strip()
                if "```" in t:
                    t = t.split("```")[1].replace("json", "", 1).strip()
                parsed = json.loads(t)
                return sanitize_pt_pt(parsed)
        except Exception as e:
            logger.warning(f"ai_json LiteLLM failed: {e}")

    # 2. Fallback direto para Gemini se disponível
    if not genai_client:
        logger.error("Nenhum provedor de LLM configurado (LiteLLM / Gemini)")
        return None
    if isinstance(model, (list, tuple)):
        model = DEFAULT_LLM_MODEL
    full_system = (system or "") + PT_PT_INSTRUCTION
    contents = prompt_or_contents if isinstance(prompt_or_contents, list) else [prompt_or_contents]
    models_to_try = [model] + [m for m in MODELS_CASCADE if m != model]
    for m in models_to_try:
        try:
            res = await genai_client.aio.models.generate_content(
                model=m,
                contents=contents,
                config=types.GenerateContentConfig(
                    system_instruction=full_system,
                    response_mime_type="application/json",
                    temperature=0.3,
                )
            )
            t = (getattr(res, "text", "") or "").strip()
            if "```" in t:
                t = t.split("```")[1].replace("json", "", 1).strip()
            if t:
                parsed = json.loads(t)
                return sanitize_pt_pt(parsed)
        except Exception as e:
            logger.warning(f"ai_json model {m} failed: {e}")
            continue
    return None

async def cached_ai(kind: str, uid: str, cid, system: str, prompt: str):
    today = datetime.now(timezone.utc).date().isoformat()
    q = {"kind": kind, "user_id": uid, "company_id": cid, "date": today}
    hit = await db.ai_cache.find_one(q)
    if hit and hit.get("payload"):
        return hit["payload"]
    payload = await ai_json(system, prompt, model=DEFAULT_LLM_MODEL)
    if payload:
        await db.ai_cache.update_one(q, {"$set": {"payload": payload, "created_at": datetime.now(timezone.utc).isoformat()}}, upsert=True)
    return payload

async def invalidate_ai_cache(uid: str):
    await db.ai_cache.delete_many({"user_id": uid})

async def _growth_score(uid: str, cid: str):
    entries = await db.entries.find({"user_id": uid, "company_id": cid}, {"type": 1, "amount": 1, "date": 1}).to_list(5000) if cid else []
    inc = {}
    for e in entries:
        mk = str(e.get("date", ""))[:7]
        if len(mk) == 7 and e["type"] == "income":
            inc[mk] = inc.get(mk, 0) + e["amount"]
    sm = sorted(inc)
    g = 50
    if len(sm) >= 2:
        recent = sum(inc[m] for m in sm[-3:]); prior = sum(inc[m] for m in sm[-6:-3])
        if prior > 0:
            g = max(5, min(100, int(60 + ((recent - prior) / prior) * 100)))
        elif recent > 0:
            g = 72
    return g

# ================================================================ FOUNDER CAMPAIGN / BILLING / ADMIN
from pymongo import ReturnDocument
from pymongo.errors import DuplicateKeyError

FOUNDER_LIMIT = 15
FOUNDER_PRICE_MONTHLY = 29
PROFESSIONAL_PRICE_MONTHLY = 59
ENTERPRISE_PRICE_MONTHLY = 159.99
PROFESSIONAL_TRIAL_DAYS = 7
FOUNDER_PROGRAM_ACTIVE_DEFAULT = True
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "").lower()

LOOKUP_TO_PLAN = {
    "founder_monthly": "founder",
    "professional_monthly": "professional",
    "premium_monthly": "professional",
    "premium_yearly": "professional",
}
PLAN_LABELS = {"founder": "Empresa Fundadora", "professional": "Professional", "enterprise": "Enterprise"}
PLAN_PRICE = {"founder": FOUNDER_PRICE_MONTHLY, "professional": PROFESSIONAL_PRICE_MONTHLY, "enterprise": ENTERPRISE_PRICE_MONTHLY}
PREMIUM_STATUSES = {"active", "trialing"}

def plan_from_lookup(lk):
    return LOOKUP_TO_PLAN.get(lk, "professional")

def is_admin_email(user: dict) -> bool:
    email = (user.get("email", "") or "").lower()
    admin_list = {"ceo@empresa.com", "d.oliveira1986@gmail.com", "obeliscolabs@gmail.com"}
    if ADMIN_EMAIL:
        admin_list.add(ADMIN_EMAIL.lower())
    return email in admin_list or user.get("role") == "admin"

async def get_admin_user(user: dict = Depends(get_current_user)) -> dict:
    if not is_admin_email(user):
        raise HTTPException(status_code=403, detail="Acesso restrito ao administrador")
    return user

async def can_access_premium(user: dict) -> bool:
    if is_admin_email(user):
        return True
    return bool(user.get("is_premium"))

async def premium_user(user: dict = Depends(get_current_user)) -> dict:
    if not (is_admin_email(user) or bool(user.get("is_premium"))):
        raise HTTPException(status_code=402, detail="premium_required")
    return user

# ---------------------------------------------------------------- campaign config
async def get_campaign() -> dict:
    c = await db.app_config.find_one({"_id": "founder_campaign"})
    if not c:
        c = {"_id": "founder_campaign", "active": FOUNDER_PROGRAM_ACTIVE_DEFAULT, "milestones_sent": []}
        try:
            await db.app_config.insert_one(c)
        except DuplicateKeyError:
            c = await db.app_config.find_one({"_id": "founder_campaign"})
    return c

async def set_campaign_active(value: bool):
    await db.app_config.update_one({"_id": "founder_campaign"}, {"$set": {"active": bool(value)}}, upsert=True)

async def founder_claimed_count() -> int:
    doc = await db.counters.find_one({"_id": "founder"})
    return int((doc or {}).get("seq", 0))

async def _allocate_founder_number():
    doc = await db.counters.find_one_and_update(
        {"_id": "founder", "seq": {"$lt": FOUNDER_LIMIT}},
        {"$inc": {"seq": 1}},
        return_document=ReturnDocument.AFTER,
    )
    return int(doc["seq"]) if doc else None

async def handle_founder_activation(user_doc: dict):
    """Atomic, race-safe founder slot allocation. Returns founder_number or None."""
    if not user_doc:
        return None
    oid = user_doc["_id"]
    if user_doc.get("founder_number") or user_doc.get("is_founder"):
        return None  # already a founder (historical) — never reallocate
    camp = await get_campaign()
    if not camp.get("active", True):
        return None
    # per-user lock: only one concurrent activation can claim
    claim = await db.users.update_one(
        {"_id": oid, "founder_number": {"$exists": False}, "is_founder": {"$ne": True},
         "founder_claim_in_progress": {"$ne": True}},
        {"$set": {"founder_claim_in_progress": True}})
    if claim.modified_count != 1:
        return None
    num = await _allocate_founder_number()
    if not num:
        await db.users.update_one({"_id": oid}, {"$unset": {"founder_claim_in_progress": ""}})
        await set_campaign_active(False)
        return None
    now = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"_id": oid}, {
        "$set": {"is_founder": True, "founder_number": num, "founder_activated_at": now,
                 "founder_price_locked": True, "founder_subscription_status": "active"},
        "$unset": {"founder_claim_in_progress": ""}})
    remaining = FOUNDER_LIMIT - num
    try:
        await notify_founder_activated(user_doc, num, remaining)
        await check_founder_milestones(remaining)
    except Exception as e:
        logger.error(f"founder notify error: {e}")
    if num >= FOUNDER_LIMIT:
        await set_campaign_active(False)
    return num

# ---------------------------------------------------------------- stripe subscription sync
def _sub_period_end(sub, item):
    return sub.get("current_period_end") or (item.get("current_period_end") if item else None)

async def sync_subscription(sub_id: str, user_id: str = None):
    if not sub_id:
        return
    try:
        sub = stripe.Subscription.retrieve(sub_id, expand=["items.data.price"])
    except Exception as e:
        logger.error(f"sync_subscription retrieve error: {e}")
        return
    items = sub.get("items", {}).get("data", [])
    item = items[0] if items else None
    price = item["price"] if item else {}
    lk = price.get("lookup_key")
    status = sub.get("status")
    cpe = _sub_period_end(sub, item)
    customer = sub.get("customer")
    md = sub.get("metadata") or {}
    uid = user_id or md.get("user_id")
    user_doc = None
    if uid:
        try:
            user_doc = await db.users.find_one({"_id": ObjectId(uid)})
        except Exception:
            user_doc = None
    if not user_doc and customer:
        user_doc = await db.users.find_one({"stripe_customer_id": customer})
    if not user_doc and sub_id:
        user_doc = await db.users.find_one({"stripe_subscription_id": sub_id})
    if not user_doc:
        logger.error(f"sync_subscription: no user for {sub_id}")
        return
    plan = plan_from_lookup(lk)
    premium = status in PREMIUM_STATUSES
    upd = {"stripe_customer_id": customer, "stripe_subscription_id": sub_id,
           "subscription_status": status, "plan": plan, "is_premium": premium,
           "current_period_end": cpe, "subscription_lookup_key": lk,
           "cancel_at_period_end": bool(sub.get("cancel_at_period_end"))}
    if premium and not user_doc.get("subscription_started_at"):
        upd["subscription_started_at"] = datetime.now(timezone.utc).isoformat()
    if status in ("canceled", "unpaid", "incomplete_expired"):
        upd["is_premium"] = False
        upd["subscription_cancelled_at"] = datetime.now(timezone.utc).isoformat()
        if user_doc.get("is_founder"):
            upd["founder_price_locked"] = False
            upd["founder_subscription_status"] = "cancelled"
    await db.users.update_one({"_id": user_doc["_id"]}, {"$set": upd})
    if plan == "founder" and status == "active":
        fresh = await db.users.find_one({"_id": user_doc["_id"]})
        await handle_founder_activation(fresh)

# ---------------------------------------------------------------- admin notifications
async def notify_founder_activated(user_doc: dict, num: int, remaining: int):
    company = await resolve_company(str(user_doc["_id"]))
    cname = (company or {}).get("name", "(empresa)")
    name = user_doc.get("name", ""); email = user_doc.get("email", "")
    now = datetime.now(timezone.utc)
    await db.admin_notifications.insert_one({
        "type": "founder_activated", "founder_number": num, "company": cname,
        "name": name, "email": email, "remaining": remaining, "read": False,
        "created_at": now.isoformat()})
    subject = f"Nova Empresa Fundadora ativada — vaga {num} de {FOUNDER_LIMIT}"
    html = (f"<div style='font-family:Arial,sans-serif;max-width:560px;margin:auto'>"
            f"<h2 style='color:#0b0c10'>Nova Empresa Fundadora ativada</h2>"
            f"<p>Uma nova Empresa Fundadora concluiu a subscrição.</p>"
            f"<table cellpadding='6' style='font-size:14px'>"
            f"<tr><td><b>Empresa</b></td><td>{cname}</td></tr>"
            f"<tr><td><b>Responsável</b></td><td>{name}</td></tr>"
            f"<tr><td><b>E-mail</b></td><td>{email}</td></tr>"
            f"<tr><td><b>Posição</b></td><td>{num} de {FOUNDER_LIMIT}</td></tr>"
            f"<tr><td><b>Preço</b></td><td>{FOUNDER_PRICE_MONTHLY} €/mês</td></tr>"
            f"<tr><td><b>Data e hora</b></td><td>{now.strftime('%d/%m/%Y %H:%M UTC')}</td></tr>"
            f"<tr><td><b>Vagas restantes</b></td><td>{remaining}</td></tr>"
            f"</table></div>")
    if ADMIN_EMAIL:
        await send_email_raw(ADMIN_EMAIL, subject, html)

async def check_founder_milestones(remaining: int):
    if remaining not in (5, 3, 1, 0):
        return
    camp = await get_campaign()
    sent = camp.get("milestones_sent", []) or []
    if remaining in sent:
        return
    await db.app_config.update_one({"_id": "founder_campaign"}, {"$addToSet": {"milestones_sent": remaining}}, upsert=True)
    if remaining == 0:
        subject = f"Programa Empresas Fundadoras concluído — {FOUNDER_LIMIT} de {FOUNDER_LIMIT} vagas preenchidas."
        body = "Todas as vagas de Empresa Fundadora foram preenchidas. O plano Professional continua disponível."
    else:
        subject = f"Programa Empresas Fundadoras — {'resta' if remaining == 1 else 'restam'} {remaining} {'vaga' if remaining == 1 else 'vagas'}"
        body = f"Restam apenas {remaining} vagas de Empresa Fundadora."
    await db.admin_notifications.insert_one({"type": "milestone", "remaining": remaining,
                                             "read": False, "created_at": datetime.now(timezone.utc).isoformat(), "text": subject})
    if ADMIN_EMAIL:
        html = f"<div style='font-family:Arial,sans-serif'><h2>{subject}</h2><p>{body}</p></div>"
        await send_email_raw(ADMIN_EMAIL, subject, html)

# ---------------------------------------------------------------- audit
async def audit_log(admin_email: str, action: str, target: str = None, before=None, after=None):
    await db.audit_log.insert_one({"admin": admin_email, "action": action, "target": target,
                                   "before": before, "after": after,
                                   "created_at": datetime.now(timezone.utc).isoformat()})
